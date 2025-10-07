import json
import math
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import io
from io import BytesIO
from django.http import HttpResponse, JsonResponse
from datetime import date, timedelta, datetime
from django.conf import settings
from django.db.models import Q
from django.db import models
import base64
import zipfile

import pandas as _pd
from decimal import Decimal as _Decimal
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from api.utils import validate_id
from api.inventories.models import InventoryMovement
from api.purchases.models import Purchase, PurchaseDetail
from api.payments.models import Payment, Installment, InstallmentAuditLog
from api.response_helpers import not_found_error_response


def get_translation(key: str, language: str = 'es') -> str:
    """
    Obtiene la traducción para una clave específica según el idioma seleccionado.

    Args:
        key (str): Clave de traducción
        language (str): Idioma ('es' o 'en')

    Returns:
        str: Texto traducido
    """
    translations = {
        'es': {
            # Títulos principales
            'product_movements_title': 'Movimientos de Productos por Tipo de Operación',
            'global_distribution_title': 'Distribución Global de Movimientos',
            'overdue_analysis_title': 'Análisis Avanzado de Cuotas Vencidas',
            'payment_methods_title': 'Análisis de Métodos de Pago',
            'most_sold_products_title': 'Productos Más Vendidos',
            'sales_summary_title': 'Resumen de Ventas',
            'input_vs_output_title': 'Distribución de Movimientos: Salidas por Venta vs Entradas por Compra',

            # Etiquetas de ejes y medidas
            'quantity_label': 'Cantidad',
            'total_label': 'Total',
            'amount_label': 'Monto',
            'units_label': 'Unidades',
            'revenue_label': 'Ingresos',
            'date_label': 'Fecha',
            'percentage_label': 'Porcentaje',
            'days_label': 'Días',

            # Gráficos específicos
            'aging_distribution': 'Distribución por Días de Mora',
            'monetary_impact': 'Impacto Monetario por Rango',
            'temporal_evolution': 'Evolución Temporal',
            'summary_table': 'Tabla de Referencias',
            'legend_title': 'Leyenda',
            'payment_method': 'Método de Pago',

            # Estados y categorías
            'no_data_available': 'Sin datos\npara este período',
            'no_data_single': 'No hay datos disponibles',
            'input_movements': 'Entradas',
            'output_movements': 'Salidas',
            'transfer_movements': 'Transferencias',
            'return_movements': 'Devoluciones',

            # Aging ranges
            'aging_1_30': '1-30 días',
            'aging_31_60': '31-60 días',
            'aging_61_90': '61-90 días',
            'aging_91_180': '91-180 días',
            'aging_180_plus': '+180 días',

            # Excel column headers
            'product_code': 'Código Producto',
            'product_name': 'Productos',
            'quantity': 'Cantidad',
            'reason': 'Motivo',
            'occurred_at': 'Fecha de Ocurrencia',
            'location_origin': 'Ubicación Origen',
            'location_destination': 'Ubicación Destino',
            'movements_by_location': 'Movimientos por Ubicación',

            # Subtítulos y descripciones
            'generated_on': 'Generado el',
            'total_units': 'unidades',
            'distribution_by': 'Distribución por',
            'comparison_with': 'Comparación con',
        },
        'en': {
            # Main titles
            'product_movements_title': 'Product Movements by Operation Type',
            'global_distribution_title': 'Global Distribution of Movements',
            'overdue_analysis_title': 'Advanced Analysis of Overdue Installments',
            'payment_methods_title': 'Payment Methods Analysis',
            'most_sold_products_title': 'Most Sold Products',
            'sales_summary_title': 'Sales Summary',
            'input_vs_output_title': 'Movement Distribution: Sales Out vs Purchase In',

            # Axis labels and measures
            'quantity_label': 'Quantity',
            'total_label': 'Total',
            'amount_label': 'Amount',
            'units_label': 'Units',
            'revenue_label': 'Revenue',
            'date_label': 'Date',
            'percentage_label': 'Percentage',
            'days_label': 'Days',

            # Specific graphics
            'aging_distribution': 'Distribution by Days Overdue',
            'monetary_impact': 'Monetary Impact by Range',
            'temporal_evolution': 'Temporal Evolution',
            'summary_table': 'Reference Table',
            'legend_title': 'Legend',
            'payment_method': 'Payment Method',

            # States and categories
            'no_data_available': 'No data\navailable for this period',
            'no_data_single': 'No data available',
            'input_movements': 'Inputs',
            'output_movements': 'Outputs',
            'transfer_movements': 'Transfers',
            'return_movements': 'Returns',

            # Aging ranges
            'aging_1_30': '1-30 days',
            'aging_31_60': '31-60 days',
            'aging_61_90': '61-90 days',
            'aging_91_180': '91-180 days',
            'aging_180_plus': '+180 days',

            # Excel column headers
            'product_code': 'Product Code',
            'product_name': 'Products',
            'quantity': 'Quantity',
            'reason': 'Reason',
            'occurred_at': 'Occurred At',
            'location_origin': 'Origin Location',
            'location_destination': 'Destination Location',
            'movements_by_location': 'Movements by Location',

            # Subtitles and descriptions
            'generated_on': 'Generated on',
            'total_units': 'units',
            'distribution_by': 'Distribution by',
            'comparison_with': 'Comparison with',
        }
    }

    # Normalize language
    lang = language.lower().strip()
    if lang not in ['es', 'en']:
        lang = 'es'  # Default to Spanish

    return translations[lang].get(key, translations['es'].get(key, key))


def convert_numpy_types(obj):
    """
    Convierte recursivamente tipos numpy a tipos nativos de Python para serialización JSON.

    Args:
        obj: Objeto que puede contener tipos numpy (dict, list, numpy types, etc.)

    Returns:
        Objeto con tipos nativos de Python
    """
    if isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, dict):
        return {key: convert_numpy_types(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [convert_numpy_types(item) for item in obj]
    elif isinstance(obj, tuple):
        return tuple(convert_numpy_types(item) for item in obj)
    elif isinstance(obj, pd.Series):
        return obj.astype(object).tolist()
    elif isinstance(obj, _Decimal):
        return float(obj)
    else:
        return obj


def _to_decimal_series(series):
    """Convert a pandas Series (or iterable) to a Series of Decimal values safely.

    Returns the original series if it's None. Missing or invalid values become Decimal('0').
    """
    if series is None:
        return series
    s = _pd.Series(series)

    def _to_dec(v):
        if v is None or _pd.isna(v):
            return _Decimal('0')
        if isinstance(v, _Decimal):
            return v
        try:
            return _Decimal(str(v))
        except Exception:
            return _Decimal('0')

    return s.apply(_to_dec)


def format_excel_worksheet(ws, df_filtered, column_mapping, header_color="366092", title_color="FFFFFF"):
    """
    Aplica formato profesional estándar a una hoja de cálculo de Excel.

    Args:
        ws: Worksheet de openpyxl
        df_filtered: DataFrame con los datos filtrados
        column_mapping: Diccionario con el mapeo de nombres de columnas
        header_color: Color de fondo para las cabeceras (hex sin #)
        title_color: Color del texto de las cabeceras (hex sin #)

    Returns:
        ws: Worksheet formateada con estilo profesional
    """
    df_filtered = df_filtered.rename(columns=column_mapping)

    headers = list(df_filtered.columns)
    ws.append(headers)

    header_font = Font(bold=True, color=title_color, size=12, name="Calibri")
    header_fill = PatternFill(start_color=header_color,
                              end_color=header_color, fill_type="solid")
    header_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)

    professional_border = Border(
        left=Side(style='medium', color='2F4F4F'),
        right=Side(style='medium', color='2F4F4F'),
        top=Side(style='medium', color='2F4F4F'),
        bottom=Side(style='medium', color='2F4F4F')
    )

    data_border = Border(
        left=Side(style='thin', color='B0C4DE'),
        right=Side(style='thin', color='B0C4DE'),
        top=Side(style='thin', color='B0C4DE'),
        bottom=Side(style='thin', color='B0C4DE')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = professional_border

    for row_num, r in enumerate(df_filtered.itertuples(index=False), start=2):
        for col_num, value in enumerate(r, start=1):
            # Limpiar fechas con timezone antes de escribir a Excel
            if isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            if isinstance(value, datetime) and value.tzinfo is not None:
                value = value.replace(tzinfo=None)

            cell = ws.cell(row=row_num, column=col_num, value=value)

            # Formato para números y monedas
            header_name = headers[col_num-1].lower()
            if isinstance(value, (int, float, _Decimal)):
                if any(keyword in header_name for keyword in ['id', 'identificador', 'numero']):
                    # Formato entero para IDs y números de identificación
                    cell.number_format = '0'
                    cell.font = Font(name="Calibri", size=10, color="2D3748")
                elif any(keyword in header_name for keyword in ['precio', 'total', 'monto', 'subtotal', 'amount', 'revenue']):
                    # Formato moneda profesional
                    cell.number_format = '"$"#,##0.00'
                    cell.font = Font(name="Calibri", size=10,
                                     bold=True, color="1B4332")
                elif any(keyword in header_name for keyword in ['cantidad', 'quantity', 'count']):
                    # Formato número entero
                    cell.number_format = '#,##0'
                    cell.font = Font(name="Calibri", size=10, color="2D3748")
                else:
                    # Formato número decimal
                    cell.number_format = '#,##0.00'
                    cell.font = Font(name="Calibri", size=10, color="2D3748")
            # Formato para fechas
            elif isinstance(value, (date, datetime)) or 'fecha' in header_name:
                cell.number_format = 'dd/mm/yyyy'
                cell.font = Font(name="Calibri", size=10, color="2D3748")
            # Formato para porcentajes
            elif isinstance(value, str) and '%' in str(value):
                cell.number_format = '0.00%'
                cell.font = Font(name="Calibri", size=10, color="E67E22")
            else:
                # Formato texto estándar
                cell.font = Font(name="Calibri", size=10, color="2D3748")

            # Alineación de celdas
            if isinstance(value, (int, float, _Decimal)):
                cell.alignment = Alignment(
                    horizontal="right", vertical="center")
            elif isinstance(value, (date, datetime)):
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
            elif isinstance(value, str) and len(str(value)) > 30:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center")

            # Bordes para todas las celdas
            cell.border = data_border

            # Colores alternados para filas
            if row_num % 2 == 0:
                cell.fill = PatternFill(
                    start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    # Ajuste del ancho de columnas
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)

        # Calcular ancho basado en el contenido con lógica mejorada
        max_length = 0

        # Considerar el largo del encabezado
        header_length = len(str(headers[col_num-1]))
        max_length = max(max_length, header_length)

        # Revisar el contenido de las celdas
        for cell in ws[column_letter]:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass

    # Ajuste inteligente del ancho
        # Anchos específicos según tipo de contenido
        header_name = headers[col_num-1].lower()
        if any(keyword in header_name for keyword in ['id', 'código', 'code']):
            adjusted_width = min(max(max_length + 1, 8), 15)
        elif any(keyword in header_name for keyword in ['fecha', 'date']):
            adjusted_width = 12
        elif any(keyword in header_name for keyword in ['precio', 'total', 'monto', 'subtotal', 'amount']):
            adjusted_width = min(max(max_length + 2, 12), 18)
        elif any(keyword in header_name for keyword in ['nombre', 'name', 'descripción', 'description']):
            adjusted_width = min(max(max_length + 2, 20), 45)
        else:
            # Ancho estándar mejorado
            adjusted_width = min(max(max_length + 2, 12), 35)

        ws.column_dimensions[column_letter].width = adjusted_width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    ws.freeze_panes = ws['A2']

    # Agregar línea separadora después del encabezado
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_num)
        current_border = cell.border
        # Agregar línea superior más gruesa
        enhanced_border = Border(
            left=current_border.left,
            right=current_border.right,
            top=Side(style='medium', color='2F4F4F'),
            bottom=current_border.bottom
        )
        cell.border = enhanced_border

    return ws


def format_summary_worksheet(ws, summary_data, title="Resumen", header_color="2E86AB"):
    """
    Formatea una hoja de resumen con KPIs.

    Args:
        ws: Worksheet de openpyxl
        summary_data: Lista de tuplas con (KPI, Valor)
        title: Título de la hoja
        header_color: Color de fondo para las cabeceras
    """
    # Título principal
    ws.merge_cells('A1:C1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=18, color="FFFFFF", name="Calibri")
    title_cell.fill = PatternFill(
        start_color=header_color, end_color=header_color, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar altura de la fila del título
    ws.row_dimensions[1].height = 30

    # Subtítulo con fecha
    from datetime import datetime
    ws.merge_cells('A2:C2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}"
    subtitle_cell.font = Font(size=10, color="7F8C8D",
                              name="Calibri", italic=True)
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Cabeceras de la tabla
    ws['A4'] = 'KPI'
    ws['B4'] = 'Valor'
    ws['C4'] = 'Categoría'

    header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    header_fill = PatternFill(start_color="34495E",
                              end_color="34495E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    professional_border = Border(
        left=Side(style='medium', color='2C3E50'),
        right=Side(style='medium', color='2C3E50'),
        top=Side(style='medium', color='2C3E50'),
        bottom=Side(style='medium', color='2C3E50')
    )

    for cell in [ws['A4'], ws['B4'], ws['C4']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = professional_border

    # Ajustar altura de la fila de cabeceras
    ws.row_dimensions[4].height = 25

    # Agregar datos de resumen con formato
    data_border = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7')
    )

    for row_num, (kpi, valor) in enumerate(summary_data, start=5):
        # Celda KPI con formato mejorado
        kpi_cell = ws.cell(row=row_num, column=1, value=kpi)
        kpi_cell.font = Font(bold=True, name="Calibri",
                             size=11, color="2C3E50")
        kpi_cell.alignment = Alignment(horizontal="left", vertical="center")
        kpi_cell.border = data_border

        # Celda Valor con formato inteligente
        valor_cell = ws.cell(row=row_num, column=2, value=valor)

    # Formato inteligente según tipo de valor
        if isinstance(valor, str) and valor.startswith('$'):
            valor_cell.font = Font(
                bold=True, name="Calibri", size=11, color="27AE60")
        elif isinstance(valor, (int, float)) and valor > 0:
            valor_cell.font = Font(
                bold=True, name="Calibri", size=11, color="3498DB")
        elif isinstance(valor, str) and '%' in valor:
            valor_cell.font = Font(
                bold=True, name="Calibri", size=11, color="E67E22")
        else:
            valor_cell.font = Font(name="Calibri", size=11, color="2C3E50")

        valor_cell.alignment = Alignment(horizontal="right", vertical="center")
        valor_cell.border = data_border

        # Celda Categoría (nueva columna para mejor organización)
        categoria = "Financiero" if isinstance(valor, str) and valor.startswith('$') else \
            "Porcentaje" if isinstance(valor, str) and '%' in valor else \
            "Cantidad" if isinstance(valor, (int, float)) else "General"

        categoria_cell = ws.cell(row=row_num, column=3, value=categoria)
        categoria_cell.font = Font(
            name="Calibri", size=10, color="7F8C8D", italic=True)
        categoria_cell.alignment = Alignment(
            horizontal="center", vertical="center")
        categoria_cell.border = data_border

    # Colores alternados para filas
        if row_num % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=row_num, column=col).fill = PatternFill(
                    start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    # Ajuste de anchos de columna
    ws.column_dimensions['A'].width = 35  # KPI
    ws.column_dimensions['B'].width = 20  # Valor
    ws.column_dimensions['C'].width = 15  # Categoría

    # Agregar línea de separación después de cabeceras
    for col_num in range(1, 4):
        cell = ws.cell(row=5, column=col_num)
        current_border = cell.border
        enhanced_border = Border(
            left=current_border.left,
            right=current_border.right,
            top=Side(style='medium', color='34495E'),
            bottom=current_border.bottom
        )
        cell.border = enhanced_border

    return ws

    return ws


def _configure_professional_matplotlib():
    """
    Configura parámetros globales de matplotlib usados por los gráficos.
    """
    import matplotlib as mpl

    mpl.rcParams.update({
        'figure.dpi': 300,
        'savefig.dpi': 300,
        'figure.facecolor': 'white',
        'figure.edgecolor': 'none',

        'font.family': 'sans-serif',
        'font.sans-serif': ['Calibri', 'Arial', 'DejaVu Sans', 'Liberation Sans'],
        'font.size': 11,
        'font.weight': 'normal',

        # Configuración de ejes
        'axes.facecolor': '#FAFBFC',
        'axes.edgecolor': '#85929E',
        'axes.linewidth': 1.2,
        'axes.grid': True,
        'axes.axisbelow': True,
        'axes.labelsize': 12,
        'axes.titlesize': 14,
        'axes.titleweight': 'bold',
        'axes.labelweight': 'medium',
        'axes.titlepad': 20,

        # Grilla
        'grid.color': '#BDC3C7',
        'grid.linestyle': '--',
        'grid.linewidth': 0.5,
        'grid.alpha': 0.4,

        # Ticks
        'xtick.labelsize': 10,
        'ytick.labelsize': 10,
        'xtick.color': '#566573',
        'ytick.color': '#566573',
        'xtick.direction': 'out',
        'ytick.direction': 'out',

        # Líneas y marcadores
        'lines.linewidth': 2,
        'lines.markersize': 6,
        'lines.markeredgewidth': 0.5,

        # Leyendas
        'legend.fancybox': True,
        'legend.framealpha': 0.9,
        'legend.shadow': True,
        'legend.frameon': True,
        'legend.edgecolor': '#E8E8E8',
        'legend.facecolor': 'white',
        'legend.fontsize': 10,

        # Guardado
        'savefig.bbox': 'tight',
        'savefig.pad_inches': 0.1,
        'savefig.facecolor': 'white',
        'savefig.edgecolor': 'none',
        'savefig.format': 'png',

        # Layout
        'figure.constrained_layout.use': True,
        'figure.constrained_layout.h_pad': 0.04167,
        'figure.constrained_layout.w_pad': 0.04167,
    })


def _get_professional_color_palette(n_colors=10):
    """
    Genera una paleta de colores profesional para visualizaciones empresariales.

    Args:
        n_colors (int): Número de colores necesarios en la paleta

    Returns:
        list: Lista de colores en formato hexadecimal profesional
    """
    base_palette = [
        '#2E86AB',
        '#A23B72',
        '#F18F01',
        '#C73E1D',
        '#2ECC71',
        '#9B59B6',
        '#F39C12',
        '#1ABC9C',
        '#34495E',
        '#E67E22',
        '#3498DB',
        '#E74C3C',
        '#95A5A6',
        '#8E44AD',
        '#16A085'
    ]

    if n_colors <= len(base_palette):
        return base_palette[:n_colors]
    else:
        extended_palette = base_palette.copy()
        from matplotlib.colors import hex2color, rgb2hex

        for i in range(n_colors - len(base_palette)):
            base_color = base_palette[i % len(base_palette)]
            rgb = hex2color(base_color)

            # Crear variación más clara o más oscura
            factor = 0.7 if i % 2 == 0 else 1.3
            new_rgb = tuple(min(1.0, max(0.0, c * factor)) for c in rgb)
            # Asegurar que tenemos exactamente 3 valores para RGB
            if len(new_rgb) == 3:
                extended_palette.append(rgb2hex(new_rgb))
            else:
                # Fallback: usar color base si hay problemas
                extended_palette.append(base_color)

        return extended_palette


def _optimize_bar_chart(ax, data, color, title, y_label, bar_subtitle, max_items=20, language='es'):
    """
    Función auxiliar para optimizar gráficos de barras con muchos elementos - Versión Profesional.

    Args:
        ax: Axes de matplotlib
        data: DataFrame con columnas 'product__name' y 'quantity'
        color: Color principal para las barras
        title: Título del gráfico
        y_label: Etiqueta del eje Y
        bar_subtitle: Subtítulo con total
        max_items: Número máximo de elementos a mostrar

    Returns:
        float: Total de cantidad procesada
    """
    if data.empty or 'quantity' not in data.columns:
        # Estado vacío
        ax.bar([], [], color=color)
        no_data_text = get_translation('no_data_available', language)
        ax.text(0.5, 0.5, no_data_text,
                transform=ax.transAxes, ha='center', va='center',
                fontsize=14, alpha=0.7, style='italic', color='#7F8C8D',
                bbox=dict(boxstyle="round,pad=0.4", facecolor='#ECF0F1', alpha=0.6, edgecolor='#BDC3C7'))
        total_quantity = 0
    else:
        # Limitar a elementos más relevantes
        top_data = data.head(max_items)

        # Truncar nombres para etiquetas
        def truncate_name(name, max_length=18):
            return str(name)[:max_length] + "..." if len(str(name)) > max_length else str(name)

        short_names = [truncate_name(str(name))
                       for name in top_data['product__name']]

    # Crear gradiente de colores
        n_bars = len(top_data)
        if n_bars > 1:
            # Generar paleta degradada basada en el color principal
            from matplotlib.colors import hex2color
            base_rgb = hex2color(color if color.startswith('#') else '#3498DB')
            colors = []
            for i in range(n_bars):
                # Crear variaciones del color base
                factor = 0.3 + 0.7 * (i / (n_bars - 1))  # De 0.3 a 1.0
                new_color = tuple(min(1.0, c + (1-c) * (1-factor))
                                  for c in base_rgb)
                colors.append(new_color)
        else:
            colors = [color]

    # Crear gráfico de barras
        bars = ax.bar(range(len(top_data)), top_data['quantity'],
                      color=colors, alpha=0.9, edgecolor='white', linewidth=2,
                      capsize=5)

    # Configurar etiquetas del eje X
        ax.set_xticks(range(len(top_data)))
        ax.set_xticklabels(short_names, rotation=45, ha="right", fontsize=10,
                           fontweight='medium', color='#2C3E50')
        ax.margins(x=0.01)

        total_quantity = data['quantity'].sum()

        for i, bar in enumerate(bars):
            height = bar.get_height()
            if height > 0:

                if height >= 1000000:
                    label = f'{height/1000000:.1f}M'
                elif height >= 1000:
                    label = f'{height/1000:.1f}k'
                elif height >= 100:
                    label = f'{int(height)}'
                else:
                    label = f'{height:.1f}'

                # Posicionamiento dinámico y profesional
                y_position = height + max(1, total_quantity * 0.02)
                ax.annotate(label,
                            xy=(bar.get_x() + bar.get_width() / 2, y_position),
                            ha='center', va='bottom',
                            fontsize=9, fontweight='bold', color='#2C3E50',
                            bbox=dict(boxstyle="round,pad=0.25", facecolor='white',
                                      alpha=0.9, edgecolor='#E8E8E8', linewidth=0.5))

    units_label = get_translation('total_units', language)
    ax.set_title(f"{title}\n({bar_subtitle}: {total_quantity:,.0f} {units_label})",
                 fontsize=14, fontweight="bold", pad=20, color="#2C3E50",
                 bbox=dict(boxstyle="round,pad=0.5", facecolor='#F8F9FA', alpha=0.8))

    ax.set_ylabel(y_label, fontsize=12, fontweight='semibold',
                  color="#34495E", labelpad=10)

    ax.set_facecolor("#FAFBFC")
    ax.grid(axis='y', linestyle='-', alpha=0.2, linewidth=1, color="#D5DBDB")
    ax.grid(axis='y', linestyle='--', alpha=0.4,
            linewidth=0.5, color="#BDC3C7")
    ax.set_axisbelow(True)

    for spine_name, spine in ax.spines.items():
        if spine_name in ['top', 'right']:
            spine.set_visible(False)
        else:
            spine.set_linewidth(1.2)
            spine.set_color('#85929E')
            spine.set_alpha(0.8)

    if total_quantity > 0:
        ax.set_ylim(0, total_quantity * 1.12)

    # Personalizar ticks del eje Y
    ax.tick_params(axis='y', labelsize=10, colors='#566573', pad=5)
    ax.tick_params(axis='x', labelsize=10, colors='#566573', pad=5)

    return total_quantity


def _generate_general_movements_graphic(df_movements, language_graphic='es'):
    """
    Genera un gráfico general de movimientos por tipo de operación.

    Args:
        df_movements (DataFrame): DataFrame con todos los movimientos
        language_graphic (str): Idioma ('es' o 'en')

    Returns:
        bytes: Contenido del gráfico PNG
    """
    # Obtener todos los movimientos (no solo entrada/salida)
    all_movements = InventoryMovement.objects.select_related(
        "product", "from_location", "to_location"
    ).filter(
        occurred_at__date__gte=df_movements['occurred_at'].min(),
        occurred_at__date__lte=df_movements['occurred_at'].max()
    ).values(
        "product__name",
        "from_location__name",
        "to_location__name",
        "quantity",
        "reason"
    )

    df_all_movements = pd.DataFrame.from_records(list(all_movements))

    if df_all_movements.empty:
        # Crear gráfico vacío
        fig, ax = plt.subplots(figsize=(10, 6))
        no_data_text = get_translation('no_data_single', language_graphic)
        ax.text(0.5, 0.5, no_data_text,
                ha='center', va='center', transform=ax.transAxes)
        title_text = get_translation(
            'product_movements_title', language_graphic)
        ax.set_title(title_text,
                     fontsize=14, fontweight='bold', pad=20, color="#2C3E50")
        buffer = BytesIO()
        fig.savefig(buffer, format='png')
        buffer.seek(0)
        content = buffer.getvalue()
        plt.close(fig)
        return content

    reasons = [
        InventoryMovement.Reason.EXIT_SALE,
        InventoryMovement.Reason.PURCHASE_ENTRY,
        InventoryMovement.Reason.TRANSFER,
        InventoryMovement.Reason.RETURN_ENTRY,
        InventoryMovement.Reason.RETURN_OUTPUT
    ]
    n_reasons = len(reasons)

    # Traducir títulos según idioma
    if language_graphic == 'es':
        main_title = get_translation('product_movements_title', 'es')
        pie_title = get_translation('global_distribution_title', 'es')
        y_label = get_translation('quantity_label', 'es')
        bar_subtitle = get_translation('total_label', 'es')
    else:
        main_title = get_translation('product_movements_title', 'en')
        pie_title = get_translation('global_distribution_title', 'en')
        y_label = get_translation('quantity_label', 'en')
        bar_subtitle = get_translation('total_label', 'en')

    cols = 3
    rows = math.ceil(n_reasons / cols)
    fig, axs = plt.subplots(rows, cols, figsize=(20, 8*rows))
    axs = axs.flatten()
    fig.suptitle(main_title, fontsize=16,
                 fontweight="bold", color="#2C3E50", y=0.96)

    max_quantity = 20

    for i, reason in enumerate(reasons):
        df_filtered = df_all_movements[df_all_movements["reason"] == reason]

        if not df_filtered.empty:
            if reason == InventoryMovement.Reason.EXIT_SALE:
                grouped = df_filtered.groupby(["product__name", "from_location__name"], as_index=False)[
                    "quantity"].sum().sort_values(by="quantity", ascending=False)
            elif reason == InventoryMovement.Reason.PURCHASE_ENTRY:
                grouped = df_filtered.groupby(["product__name", "to_location__name"], as_index=False)[
                    "quantity"].sum().sort_values(by="quantity", ascending=False)
            elif reason == InventoryMovement.Reason.TRANSFER:
                grouped = df_filtered.groupby(["product__name", "from_location__name", "to_location__name"], as_index=False)[
                    "quantity"].sum().sort_values(by="quantity", ascending=False)
            elif reason == InventoryMovement.Reason.RETURN_ENTRY:
                grouped = df_filtered.groupby(["product__name", "to_location__name"], as_index=False)[
                    "quantity"].sum().sort_values(by="quantity", ascending=False)
            elif reason == InventoryMovement.Reason.RETURN_OUTPUT:
                grouped = df_filtered.groupby(["product__name", "from_location__name"], as_index=False)[
                    "quantity"].sum().sort_values(by="quantity", ascending=False)

            if not grouped.empty and 'quantity' in grouped.columns:
                max_quantity = max(max_quantity, grouped["quantity"].max() + 5)
                colors = ["#3498db", "#e74c3c", "#2ecc71",
                          "#9b59b6", "#f39c12", "#1abc9c"]
                grouped.plot(
                    kind="bar", ax=axs[i], color=colors[i % len(colors)])
                axs[i].set_xticklabels(
                    grouped.product__name, rotation=90, ha="right")
                total_quantity = grouped['quantity'].sum()
            else:
                # Crear gráfico vacío
                axs[i].bar([], [], color=colors[i % len(colors)]
                           if 'colors' in locals() else '#3498db')
                axs[i].set_xticklabels([])
                total_quantity = 0
        else:
            # Crear gráfico vacío
            colors = ["#3498db", "#e74c3c", "#2ecc71",
                      "#9b59b6", "#f39c12", "#1abc9c"]
            axs[i].bar([], [], color=colors[i % len(colors)])
            axs[i].set_xticklabels([])
            total_quantity = 0

        axs[i].set_title(
            f"{str(reason).replace('_', ' ').title()} ({bar_subtitle}: {total_quantity})")
        axs[i].set_ylabel(y_label)
        axs[i].legend().remove()
        axs[i].set_facecolor("#DBDBDB")
        axs[i].grid(axis='y', linestyle='--', alpha=0.7)

        # Agregar anotaciones en las barras
        for p in axs[i].patches:
            try:
                height = p.get_height()
                if height > 0:
                    axs[i].annotate(str(int(height)),
                                    (p.get_x() + p.get_width() / 2., height),
                                    ha='center', va='bottom', fontsize=8, color="black")
            except:
                pass

    # Establecer límites uniformes
    for i in range(n_reasons):
        axs[i].set_ylim(0, max_quantity)

    if len(axs) > n_reasons:
        totals_by_reason = df_all_movements.groupby("reason")["quantity"].sum()
        if not totals_by_reason.empty:
            ax_pie = axs[-1]
            n_slices = len(totals_by_reason)
            explode_values = [0.1 if i == 0 else 0.05 for i in range(n_slices)]
            ax_pie.pie(
                totals_by_reason,
                labels=[str(totals_by_reason.index[i]).replace(
                    '_', ' ').title() for i in range(len(totals_by_reason))],
                autopct="%1.1f%%",
                startangle=90,
                explode=explode_values,
                shadow=True,
                colors=_get_professional_color_palette(n_slices),
                wedgeprops=dict(width=0.9, edgecolor='white', linewidth=2)
            )
            ax_pie.set_title(pie_title, fontsize=14,
                             fontweight="bold", pad=20, color="#2C3E50")

    # Ocultar subplots sobrantes
    for j in range(n_reasons, len(axs)):
        if j != len(axs)-1:  # no tocar la celda del pie
            axs[j].axis("off")

    # Ajustar layout con espacio suficiente para el título
    plt.subplots_adjust(top=1.0, bottom=0.08, left=0.08,
                        right=0.95, hspace=0.6, wspace=0.6)
    # rect deja espacio para suptitle
    plt.tight_layout(rect=(0, 0, 1, 0.90))

    # Guardar en buffer
    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
    buffer.seek(0)
    content = buffer.getvalue()
    plt.close(fig)

    return content


def product_rotation_by_location(
    location_id: int,
    graphic: bool,
    from_date: date,
    to_date: date,
    download_graphic: bool = False,
    excel: bool = False,
    language_graphic: str = 'es'
):
    """
    Reporte de rotación de productos por ubicación (entrada/salida/transferencias/retornos).

    Este servicio agrega los movimientos de inventario para una ubicación específica
    en un rango de fechas y devuelve los registros junto con un gráfico opcional
    (como imagen base64) y/o un archivo Excel descargable.

    Args:
        location_id (int): ID de la ubicación (almacén / depósito).
        graphic (bool): Si se debe generar un gráfico (True) para visualización.
        from_date (date): Fecha inicial del intervalo (inclusive).
        to_date (date): Fecha final del intervalo (inclusive).
        download_graphic (bool, optional): Si se deben empaquetar Excel + gráfico en un ZIP para descarga. Default False.
        excel (bool, optional): Si se debe generar un archivo Excel con los datos. Default False.
        language_graphic (str, optional): Idioma para títulos/etiquetas ('es' o 'en'). Default 'es'.

    Returns:
        JsonResponse | HttpResponse: Si `excel`/`download_graphic` está activo devuelve un `HttpResponse`
            con el archivo correspondiente (Excel o ZIP). En modo normal devuelve un `JsonResponse`
            con la estructura:

            {
                "success": True,
                "message": "Data retrieved successfully",
                "data": [ ... lista de movimientos ... ],
                "graphic_base64": "..."  # base64 PNG o None
            }

    Raises:
        ValueError: Si las fechas son inválidas (from_date > to_date, fechas en futuro, etc.).

    Nota:
        - Utiliza `InventoryMovement.Reason` para agrupar por tipo de movimiento.
        - Para la generación del gráfico se usa `matplotlib` y se expone como base64 para el front-end.
    """

    if from_date >= to_date:
        raise ValueError("from_date debe ser anterior a to_date")
    if to_date > date.today():
        raise ValueError("to_date no puede ser mayor que la fecha actual")
    if from_date > date.today():
        raise ValueError("from_date no puede ser mayor que la fecha actual")
    if from_date < date.today()-timedelta(days=365*3):
        raise ValueError("from_date must be within the last 3 years")
    validate_id(location_id, "Location")

    movements = InventoryMovement.objects.select_related(
        "product", "from_location", "to_location"
    ).filter(
        Q(from_location_id=location_id) | Q(to_location_id=location_id),
        occurred_at__date__gte=from_date,
        occurred_at__date__lte=to_date
    ).order_by('-product')

    if not movements.exists():
        # Verificar si hay movimientos para esta ubicación
        total_movements_for_location = InventoryMovement.objects.filter(
            Q(from_location_id=location_id) | Q(to_location_id=location_id)
        ).count()

        total_movements = InventoryMovement.objects.count()

        if total_movements == 0:
            return JsonResponse({
                "success": False,
                "message": "No se encontraron movimientos de inventario en la base de datos.",
                "data": {
                    "error_type": "no_inventory_movements",
                    "location_id": location_id,
                    "total_movements_in_db": 0,
                    "requested_range": {
                        "from_date": from_date.strftime("%Y-%m-%d"),
                        "to_date": to_date.strftime("%Y-%m-%d")
                    },
                    "suggestions": [
                        "Realizar operaciones de inventario en esta ubicación",
                        "Verificar que la ubicación existe y esté activa",
                        "Crear productos y realizar movimientos de inventario"
                    ]
                }
            }, status=404)
        elif total_movements_for_location == 0:
            return JsonResponse({
                "success": False,
                "message": f"No se encontraron movimientos de inventario para la ubicación ID {location_id}.",
                "data": {
                    "error_type": "no_movements_for_location",
                    "location_id": location_id,
                    "total_movements_in_db": total_movements,
                    "movements_for_location": 0,
                    "requested_range": {
                        "from_date": from_date.strftime("%Y-%m-%d"),
                        "to_date": to_date.strftime("%Y-%m-%d")
                    },
                    "suggestions": [
                        f"Realizar operaciones de inventario en la ubicación ID {location_id}",
                        "Verificar que el ID de ubicación es correcto",
                        "Crear movimientos de inventario en esta ubicación"
                    ]
                }
            }, status=404)
        else:
            # Hay movimientos para la ubicación pero no en el rango
            oldest_movement = InventoryMovement.objects.filter(
                Q(from_location_id=location_id) | Q(to_location_id=location_id)
            ).order_by('occurred_at').first()
            newest_movement = InventoryMovement.objects.filter(
                Q(from_location_id=location_id) | Q(to_location_id=location_id)
            ).order_by('-occurred_at').first()

            return JsonResponse({
                "success": False,
                "message": f"No se encontraron movimientos de inventario para la ubicación ID {location_id} en el rango especificado ({from_date.strftime('%Y-%m-%d')} a {to_date.strftime('%Y-%m-%d')}).",
                "data": {
                    "error_type": "no_movements_in_range_for_location",
                    "location_id": location_id,
                    "total_movements_for_location": total_movements_for_location,
                    "requested_range": {
                        "from_date": from_date.strftime("%Y-%m-%d"),
                        "to_date": to_date.strftime("%Y-%m-%d")
                    },
                    "available_range": {
                        "oldest_movement": oldest_movement.occurred_at.date().strftime("%Y-%m-%d") if oldest_movement else None,
                        "newest_movement": newest_movement.occurred_at.date().strftime("%Y-%m-%d") if newest_movement else None
                    },
                    "suggestions": [
                        f"Ajustar el rango de fechas entre {oldest_movement.occurred_at.date().strftime('%Y-%m-%d') if oldest_movement else 'N/A'} y {newest_movement.occurred_at.date().strftime('%Y-%m-%d') if newest_movement else 'N/A'}",
                        f"Realizar más operaciones en la ubicación ID {location_id} durante el rango deseado"
                    ]
                }
            }, status=404)
    value_movements = list(movements.values(
        "id",
        "product__product_code",
        "product__name",
        "from_location__id",
        "from_location__name",
        "to_location__id",
        "to_location__name",
        "quantity",
        "reason",
        "reference_type",
        "reference_id",
        "occurred_at"
    ))
    df_movements = pd.DataFrame.from_records(value_movements)
    df_movements['occurred_at'] = pd.to_datetime(
        df_movements['occurred_at']).dt.date

    # Mapeo de columnas para nombres más legibles (lo necesitamos antes para reutilizarlo)
    column_mapping = {
        'id': 'ID',
        'product__product_code': get_translation('product_code', language_graphic),
        'product__name': get_translation('product_name', language_graphic),
        'from_location__id': f"ID {get_translation('location_origin', language_graphic)}",
        'from_location__name': get_translation('location_origin', language_graphic),
        'to_location__id': f"ID {get_translation('location_destination', language_graphic)}",
        'to_location__name': get_translation('location_destination', language_graphic),
        'quantity': get_translation('quantity', language_graphic),
        'reason': get_translation('reason', language_graphic),
        'reference_type': 'Tipo Referencia' if language_graphic == 'es' else 'Reference Type',
        'reference_id': 'ID Referencia' if language_graphic == 'es' else 'Reference ID',
        'occurred_at': get_translation('occurred_at', language_graphic)
    }

    graphic_file = None
    graphic_data_for_front = None
    graphic_content = None
    if graphic:
        # Configurar matplotlib para gráficos
        _configure_professional_matplotlib()

        reasons = [
            InventoryMovement.Reason.EXIT_SALE,
            InventoryMovement.Reason.PURCHASE_ENTRY,
            InventoryMovement.Reason.TRANSFER,
            InventoryMovement.Reason.RETURN_ENTRY,
            InventoryMovement.Reason.RETURN_OUTPUT,
        ]
        n_reasons = len(reasons)

        language = language_graphic.lower().strip()

        main_title = get_translation('product_movements_title', language)
        pie_title = get_translation('global_distribution_title', language)
        y_label = get_translation('quantity_label', language)
        bar_subtitle = get_translation('total_label', language)

    # Configuración del layout
        cols = 3
        rows = math.ceil(n_reasons / cols)
        # Dimensiones optimizadas para visualización profesional
        fig, axs = plt.subplots(rows, cols, figsize=(26, 8*rows))
        axs = axs.flatten()

        fig.suptitle(main_title, fontsize=16, fontweight="bold",
                     y=0.96, color="#2C3E50",
                     bbox=dict(boxstyle="round,pad=0.6", facecolor='#ECF0F1', alpha=0.8))

        colors_palette = _get_professional_color_palette(n_reasons)

        # initialize gro`uped to a default empty DataFrame to avoid "possibly unbound" issues
        grouped = pd.DataFrame(
            columns=["product__name", "quantity", "from_location__name", "to_location__name"])

        for i, reason in enumerate(reasons):
            df_filtered = df_movements[df_movements["reason"] == reason]

            # Preparar datos según tipo de movimiento
            if reason == InventoryMovement.Reason.EXIT_SALE:
                if not df_filtered.empty:
                    grouped = df_filtered.groupby(["product__name", "from_location__name"], as_index=False)[
                        "quantity"].sum().sort_values(by="quantity", ascending=False)
                else:
                    grouped = pd.DataFrame(
                        columns=["product__name", "quantity", "from_location__name"])
            elif reason == InventoryMovement.Reason.PURCHASE_ENTRY:
                if not df_filtered.empty:
                    grouped = df_filtered.groupby(["product__name", "to_location__name"], as_index=False)[
                        "quantity"].sum().sort_values(by="quantity", ascending=False)
                else:
                    grouped = pd.DataFrame(
                        columns=["product__name", "quantity", "to_location__name"])
            elif reason == InventoryMovement.Reason.TRANSFER:
                if not df_filtered.empty:
                    grouped = df_filtered.groupby(["product__name", "from_location__name", "to_location__name"], as_index=False)[
                        "quantity"].sum().sort_values(by="quantity", ascending=False)
                else:
                    grouped = pd.DataFrame(
                        columns=["product__name", "quantity", "from_location__name", "to_location__name"])
            elif reason == InventoryMovement.Reason.RETURN_ENTRY:
                if not df_filtered.empty:
                    grouped = df_filtered.groupby(["product__name", "to_location__name"], as_index=False)[
                        "quantity"].sum().sort_values(by="quantity", ascending=False)
                else:
                    grouped = pd.DataFrame(
                        columns=["product__name", "quantity", "to_location__name"])
            elif reason == InventoryMovement.Reason.RETURN_OUTPUT:
                if not df_filtered.empty:
                    grouped = df_filtered.groupby(["product__name", "from_location__name"], as_index=False)[
                        "quantity"].sum().sort_values(by="quantity", ascending=False)
                else:
                    grouped = pd.DataFrame(
                        columns=["product__name", "quantity", "from_location__name"])

            # Generar gráfico con función optimizada
            color = colors_palette[i % len(colors_palette)]
            reason_title = str(reason).replace('_', ' ').title()

            # Usar la función optimizada para crear gráficos profesionales
            total_quantity = _optimize_bar_chart(
                ax=axs[i],
                data=grouped,
                color=color,
                title=reason_title,
                y_label=y_label,
                bar_subtitle=bar_subtitle,
                max_items=15,
                language=language
            )

    # Gráfico de pastel en la última celda
        if len(axs) > n_reasons:  # verificamos que haya espacio libre
            totals_by_reason = df_movements.groupby("reason")["quantity"].sum()
            if not totals_by_reason.empty:
                ax_pie = axs[-1]

                # Colores consistentes
                pie_colors = colors_palette[:len(totals_by_reason)]

                # Crear etiquetas con porcentajes para la leyenda
                total = totals_by_reason.sum()
                labels_with_percentage = []
                for reason, value in totals_by_reason.items():
                    percentage = (value / total * 100) if total > 0 else 0
                    clean_reason = str(reason).replace('_', ' ').title()
                    labels_with_percentage.append(
                        f'{clean_reason}: {percentage:.1f}% ({value:,})')

                wedges, texts = ax_pie.pie(
                    totals_by_reason,
                    labels=None,  # Sin etiquetas en el gráfico
                    autopct=None,  # Sin porcentajes en el gráfico
                    startangle=90,
                    colors=pie_colors,
                    # Separación sutil uniforme
                    explode=[0.02] * len(totals_by_reason),
                    shadow=True,
                    wedgeprops=dict(width=0.9, edgecolor='white', linewidth=2),
                    textprops={'fontsize': 9, 'fontweight': 'medium'}
                )

                legend = ax_pie.legend(
                    wedges,
                    labels_with_percentage,
                    loc="center left",
                    bbox_to_anchor=(1.05, 0, 0.5, 1),
                    fontsize=9,
                    title_fontsize=10,
                    frameon=True,
                    fancybox=True,
                    shadow=True,
                    framealpha=0.95,
                    edgecolor='#BDC3C7',
                    facecolor='white'
                )
                legend.get_title().set_fontweight('bold')

                # Título
                ax_pie.set_title(pie_title,
                                 fontsize=14, fontweight="bold", pad=20, color="#2C3E50",
                                 bbox=dict(boxstyle="round,pad=0.5", facecolor='#F8F9FA', alpha=0.8))

                # Mejorar la visualización del círculo
                ax_pie.axis('equal')
            else:
                # Si no hay datos, crear mensaje
                axs[-1].text(0.5, 0.5, 'Sin datos suficientes\npara distribución',
                             transform=axs[-1].transAxes, ha='center', va='center',
                             fontsize=12, alpha=0.7, style='italic', color='#7F8C8D',
                             bbox=dict(boxstyle="round,pad=0.4", facecolor='#ECF0F1', alpha=0.6))
                axs[-1].axis("off")

        for j in range(n_reasons, len(axs)):
            if j != len(axs)-1:
                axs[j].axis("off")
                axs[j].text(0.5, 0.5, 'Reservado para\nfuturos análisis',
                            transform=axs[j].transAxes, ha='center', va='center',
                            fontsize=10, alpha=0.3, style='italic', color='#95A5A6')

    # Ajustar layout
        plt.subplots_adjust(top=0.93, bottom=0.08, left=0.08, right=0.95,
                            hspace=0.4, wspace=0.3)
        fig.tight_layout(pad=5.0)

        from datetime import datetime
        fig.text(
            0.99,
            0.01,
            f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")}',
            ha='right',
            va='bottom',
            fontsize=8,
            alpha=0.6,
            style='italic',
            color='#7F8C8D'
        )
        graphic_buffer = BytesIO()
        fig.savefig(
            graphic_buffer,
            format='png',
            dpi=350,
            bbox_inches='tight',
            facecolor='white',
            edgecolor='none',
            pad_inches=0.2,
            metadata={'Title': main_title, 'Author': 'Sistema Analytics Pro'}
        )
        graphic_buffer.seek(0)
        graphic_file = graphic_buffer
        graphic_content = graphic_buffer.getvalue()

        import base64
        graphic_data_for_front = base64.b64encode(
            graphic_content).decode('utf-8')

        plt.close(fig)

    # Manejo de respuestas según parámetros
    # Comportamiento esperado (coincidente con products_movements_input_vs_output):
    # - excel + graphic: generar Excel con gráfico embebido en hoja 2
    # - excel + graphic + download_graphic: generar ZIP con Excel + PNG
    # - excel (sin graphic): generar solo Excel
    # - graphic + download_graphic (sin excel): generar PNG
    # - otro: devolver JSON con gráfico en base64

    # Si el usuario solicitó gráfico pero no se generó contenido, crear un placeholder
    if graphic and not graphic_content:
        try:
            fig_placeholder, ax_placeholder = plt.subplots(figsize=(8, 4))
            ax_placeholder.text(0.5, 0.5, 'No graphic available',
                                ha='center', va='center', fontsize=14, color='#7F8C8D')
            ax_placeholder.axis('off')
            buf_ph = BytesIO()
            fig_placeholder.savefig(
                buf_ph, format='png', dpi=150, bbox_inches='tight')
            buf_ph.seek(0)
            graphic_content = buf_ph.getvalue()
            graphic_data_for_front = base64.b64encode(
                graphic_content).decode('utf-8')
            plt.close(fig_placeholder)
        except Exception:
            # si falla la generación del placeholder, seguimos sin gráfico pero no rompemos
            graphic_content = None

    if excel and graphic:
        # Generar Excel y embeber gráfico en segunda hoja (si existe)
        excel_buffer = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = get_translation('movements_by_location', language_graphic)

        # Aplicar formato usando la función auxiliar
        ws = format_excel_worksheet(
            ws, df_movements, column_mapping, header_color="3498DB")

        # Agregar segunda hoja con gráfico si hay contenido
        graphic_title = 'Gráfico' if language_graphic == 'es' else 'Graphic'
        ws_graphic = wb.create_sheet(title=graphic_title)
        if graphic_content:
            img_buffer = BytesIO(graphic_content)
            img = Image(img_buffer)
            # Tamaño proporcional para Excel
            img.width = 800
            img.height = 600
            ws_graphic.add_image(img, 'A1')
        else:
            # Inserta un mensaje si no hay gráfico disponible
            ws_graphic['A1'] = 'No graphic available for the selected range/parameters.'

        wb.save(excel_buffer)
        excel_buffer.seek(0)

        if download_graphic:
            # ZIP con Excel (con gráfico embebido) + PNG separado (si existe)
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w') as zipf:
                zipf.writestr("movements.xlsx", excel_buffer.getvalue())
                if graphic_content:
                    zipf.writestr("graphic.png", graphic_content)
            zip_buffer.seek(0)
            resp = HttpResponse(zip_buffer.getvalue(),
                                content_type='application/zip')
            resp['Content-Disposition'] = 'attachment; filename="product_rotation_report.zip"'
            return resp
        else:
            # Solo Excel con gráfico embebido
            resp = HttpResponse(excel_buffer.getvalue(
            ), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = 'attachment; filename="movements.xlsx"'
            return resp

    elif excel and not graphic:
        # Solo Excel sin gráfico
        excel_buffer = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = get_translation('movements_by_location', language_graphic)

        # Aplicar formato usando la función auxiliar
        ws = format_excel_worksheet(
            ws, df_movements, column_mapping, header_color="3498DB")

        wb.save(excel_buffer)
        excel_buffer.seek(0)

        resp = HttpResponse(excel_buffer.getvalue(
        ), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="movements.xlsx"'
        return resp

    elif graphic and download_graphic:
        # Solo PNG para descarga (solo si graphic=True)
        if not graphic_content:
            # intentar generar placeholder nuevamente
            try:
                fig_placeholder, ax_placeholder = plt.subplots(figsize=(8, 4))
                ax_placeholder.text(0.5, 0.5, 'No graphic available',
                                    ha='center', va='center', fontsize=14, color='#7F8C8D')
                ax_placeholder.axis('off')
                buf_ph = BytesIO()
                fig_placeholder.savefig(
                    buf_ph, format='png', dpi=150, bbox_inches='tight')
                buf_ph.seek(0)
                graphic_content = buf_ph.getvalue()
                graphic_data_for_front = base64.b64encode(
                    graphic_content).decode('utf-8')
                plt.close(fig_placeholder)
            except Exception:
                graphic_content = None

        if graphic_content:
            resp = HttpResponse(graphic_content, content_type='image/png')
            resp['Content-Disposition'] = 'attachment; filename="product_rotation_graphic.png"'
            return resp

    # Respuesta JSON con datos y gráfico en base64 para el front
    response_data = {
        "success": True,
        "message": "Data retrieved successfully",
        "data": df_movements.to_dict(orient="records"),
        "graphic_base64": graphic_data_for_front
    }
    return JsonResponse(response_data)


# Servicio para entrada y salida de productos (filtro por fecha)
def products_movements_input_vs_output(
    from_date: date,
    to_date: date,
    download_graphic: bool = False,
    excel: bool = False,
    graphic: bool = True,
    type_graphic: str = 'pie',
    language_graphic: str = 'es',
):
    """
    Servicio para obtener movimientos de entrada y salida de productos o gráfico general.

    Con los nuevos requisitos implementados:
    1. Excel + graphic = incluir gráfico en segunda hoja
    2. download_graphic solo funciona si graphic=True
    3. Excel + graphic + download_graphic = ZIP con ambos archivos
    4. Respuestas JSON incluyen gráfico en base64 para front-end
    5. type_graphic='general' muestra todos los movimientos entre fechas

    Args:
        from_date (date): Fecha de inicio
        to_date (date): Fecha final
        download_graphic (bool): Descargar PNG solo si graphic=True
        excel (bool): Generar archivo Excel
        graphic (bool): Generar gráfico
        type_graphic (str): 'pie', 'bar' o 'general' para todos los movimientos
        language_graphic (str): 'es' o 'en'

    Returns:
        HttpResponse o JsonResponse: Según parámetros, puede retornar Excel, PNG, ZIP o JSON
    """
    if from_date >= to_date:
        raise ValueError("from_date debe ser anterior a to_date")
    if to_date > date.today():
        raise ValueError("to_date no puede ser mayor que la fecha actual")
    if from_date > date.today():
        raise ValueError("from_date no puede ser mayor que la fecha actual")
    if from_date < date.today()-timedelta(days=365*3):
        raise ValueError("from_date must be within the last 3 years")
    if not isinstance(type_graphic, str):
        raise ValueError("type_graphic must be a string")

    # normalize string input
    type_graphic = type_graphic.lower().strip()

    if type_graphic not in ['pie', 'bar', 'general']:
        raise ValueError("type_graphic must be 'pie', 'bar' or 'general'")

    # Si type_graphic es 'general', usar el nuevo gráfico
    if type_graphic == 'general':
        movements = InventoryMovement.objects.select_related(
            "product", "from_location", "to_location"
        ).filter(
            occurred_at__date__gte=from_date,
            occurred_at__date__lte=to_date
        ).order_by('-product')

        if not movements.exists():
            return JsonResponse({
                "success": False,
                "message": "No se encontraron movimientos en el rango especificado.",
                "data": {
                    "error_type": "no_movements_in_range",
                    "requested_range": {
                        "from_date": from_date.strftime("%Y-%m-%d"),
                        "to_date": to_date.strftime("%Y-%m-%d")
                    }
                }
            }, status=404)

        # Crear DataFrame básico para Excel
        value_movements = movements.values(
            "id", "product__product_code", "product__name", "quantity",
            "reason", "occurred_at"
        )
        df_movements = pd.DataFrame.from_records(list(value_movements))
        df_movements['occurred_at'] = pd.to_datetime(
            df_movements['occurred_at']).dt.date

        # Generar gráfico general
        graphic_content = None
        graphic_data_for_front = None

        if graphic:
            graphic_content = _generate_general_movements_graphic(
                df_movements, language_graphic)

            if graphic_content:
                graphic_data_for_front = base64.b64encode(
                    graphic_content).decode('utf-8')

        # Manejo de respuestas según parámetros
        if excel and graphic and graphic_content:
            # Crear Excel con gráfico en segunda hoja
            excel_buffer = BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws:
                ws.title = 'Movimientos Generales'

                column_mapping = {
                    'id': 'ID', 'product__product_code': 'Código Producto',
                    'product__name': 'Nombre Producto', 'quantity': 'Cantidad',
                    'reason': 'Motivo', 'occurred_at': 'Fecha'
                }
                ws = format_excel_worksheet(
                    ws, df_movements, column_mapping, header_color="E74C3C")

                # Agregar segunda hoja con gráfico proporcional
                ws_graphic = wb.create_sheet(title="Gráfico")
                img_buffer = BytesIO(graphic_content)
                img = Image(img_buffer)
                # Tamaño proporcional para Excel
                img.width = 480
                img.height = 320
                ws_graphic.add_image(img, 'A1')

                wb.save(excel_buffer)
                excel_buffer.seek(0)

                if download_graphic:
                    # ZIP con Excel y PNG
                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w') as zipf:
                        zipf.writestr("movimientos_generales.xlsx",
                                      excel_buffer.getvalue())
                        zipf.writestr("grafico_general.png", graphic_content)
                    zip_buffer.seek(0)
                    resp = HttpResponse(
                        zip_buffer.getvalue(), content_type='application/zip')
                    resp['Content-Disposition'] = 'attachment; filename="movimientos_generales.zip"'
                    return resp
                else:
                    # Solo Excel con gráfico
                    resp = HttpResponse(excel_buffer.getvalue(
                    ), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    resp['Content-Disposition'] = 'attachment; filename="movimientos_generales.xlsx"'
                    return resp

        elif excel and not graphic:
            # Solo Excel sin gráfico
            excel_buffer = BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws:
                ws.title = 'Movimientos Generales'

                column_mapping = {
                    'id': 'ID', 'product__product_code': 'Código Producto',
                    'product__name': 'Nombre Producto', 'quantity': 'Cantidad',
                    'reason': 'Motivo', 'occurred_at': 'Fecha'
                }
                ws = format_excel_worksheet(
                    ws, df_movements, column_mapping, header_color="E74C3C")

                wb.save(excel_buffer)
                excel_buffer.seek(0)
                resp = HttpResponse(excel_buffer.getvalue(
                ), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                resp['Content-Disposition'] = 'attachment; filename="movimientos_generales.xlsx"'
                return resp

        elif graphic and download_graphic and graphic_content:
            # Solo PNG para descarga
            resp = HttpResponse(graphic_content, content_type='image/png')
            resp['Content-Disposition'] = 'attachment; filename="movimientos_generales.png"'
            return resp

        else:
            # JSON con gráfico base64 para front-end
            return JsonResponse({
                "success": True,
                "message": "Movimientos generales obtenidos exitosamente",
                "data": df_movements.to_dict(orient="records"),
                "graphic_base64": graphic_data_for_front
            })

    # Lógica original para type_graphic = 'pie' o 'bar'
    movements = InventoryMovement.objects.select_related(
        "product", "from_location", "to_location"
    ).filter(
        occurred_at__date__gte=from_date,
        occurred_at__date__lte=to_date,
        reason__in=[InventoryMovement.Reason.EXIT_SALE,
                    InventoryMovement.Reason.PURCHASE_ENTRY]
    ).order_by('-product')

    if not movements.exists():
        total_movements = InventoryMovement.objects.count()
        if total_movements == 0:
            return JsonResponse({
                "success": False,
                "message": "No se encontraron movimientos de inventario en la base de datos.",
                "data": {
                    "error_type": "no_inventory_movements",
                    "total_movements_in_db": 0,
                    "requested_range": {
                        "from_date": from_date.strftime("%Y-%m-%d"),
                        "to_date": to_date.strftime("%Y-%m-%d")
                    }
                }
            }, status=404)
        else:
            return JsonResponse({
                "success": False,
                "message": f"No se encontraron movimientos de entrada/salida en el rango especificado.",
                "data": {
                    "error_type": "no_movements_in_range",
                    "requested_range": {
                        "from_date": from_date.strftime("%Y-%m-%d"),
                        "to_date": to_date.strftime("%Y-%m-%d")
                    }
                }
            }, status=404)

    value_movements = movements.values(
        "id", "product__product_code", "product__name", "product__description",
        "product__brand", "product__model", "from_location__id", "from_location__name",
        "to_location__id", "to_location__name", "quantity", "reason",
        "reference_type", "reference_id", "occurred_at"
    )

    df_movements = pd.DataFrame.from_records(list(value_movements))
    df_movements['occurred_at'] = pd.to_datetime(
        df_movements['occurred_at']).dt.date

    # Generar gráfico si se solicita
    graphic_content = None
    graphic_data_for_front = None
    fig = None

    if graphic:
        grouped = df_movements.groupby(['reason'], as_index=False)[
            'quantity'].sum()

        if grouped.empty:
            max_quantity = 50
        else:
            max_quantity = grouped['quantity'].max(
            ) + (grouped['quantity'].max()*0.1)

        main_title = get_translation('input_vs_output_title', language_graphic)
        legend_title = get_translation('legend_title', language_graphic)
        units_label = get_translation('total_units', language_graphic)

        if type_graphic == 'pie':
            fig, ax = plt.subplots(
                figsize=(10, 8), subplot_kw=dict(aspect="equal"))

            colors = ["#f57d7d", "#917bf5"]

            # Crear efecto de profundidad multicapa
            explode_values = [0.15, 0.08]  # Separación progresiva

            pie_result = ax.pie(
                grouped['quantity'],
                explode=explode_values,
                shadow=True,
                startangle=45,
                colors=colors,
                autopct='%1.1f%%',           # Mostrar porcentajes
                textprops=dict(
                    color="white",
                    fontsize=12,
                    fontweight="bold",
                    bbox=dict(                # Caja alrededor del texto para profundidad
                        boxstyle="round,pad=0.3",
                        facecolor='black',
                        alpha=0.7,
                        edgecolor='white'
                    )
                ),
                wedgeprops=dict(
                    edgecolor='#2C3E50',     # Borde oscuro para contraste
                    linewidth=3,             # Borde grueso para definición
                    antialiased=True,
                    alpha=0.95               # Ligera transparencia para suavidad
                ),
                # Configuración de profundidad
                radius=1.1,                   # Radio aumentado
                pctdistance=0.8,             # Porcentajes más cerca del centro
                labeldistance=1.15,          # Labels más alejados
                rotatelabels=True,           # Rotar labels para mejor lectura
                normalize=True,              # Normalizar para precisión
                counterclock=False           # Sentido horario para convención
            )

            # Handle pie result unpacking
            if len(pie_result) >= 2:
                wedges = pie_result[0]
            else:
                wedges = pie_result

            # Create legend with percentages
            total_quantity = grouped['quantity'].sum()
            legend_labels = []
            for i, (reason, quantity) in enumerate(zip(grouped['reason'], grouped['quantity'])):
                percentage = (quantity / total_quantity) * 100
                clean_reason = reason.replace('_', ' ').title()
                legend_labels.append(
                    f"{clean_reason}: {percentage:.1f}% ({quantity:,} {units_label})")

            ax.legend(wedges, legend_labels,
                      title=legend_title, loc="center left", bbox_to_anchor=(1, 0, 0.5, 1), fontsize=11)

            # Título más compacto y mejor posicionado
            if language_graphic.lower().strip() == 'en':
                compact_title = "Movement Distribution:\nSales Out vs Purchase In"
            else:
                compact_title = "Distribución de Movimientos:\nSalidas por Venta vs Entradas por Compra"

            ax.set_title(compact_title, fontsize=22, fontweight="bold",
                         y=0.96, color="#2C3E50",
                         bbox=dict(boxstyle="round,pad=0.6", facecolor='#ECF0F1', alpha=0.8))

        elif type_graphic == 'bar':
            fig, ax = plt.subplots(figsize=(12, 10))
            colors = ["#f57d7d", "#917bf5"]
            bar_labels = [f"{grouped['reason'].iloc[i].replace('_', ' ').title()} ({grouped['quantity'].iloc[i]})"
                          for i in range(len(grouped))]

            bars = ax.bar(grouped['reason'].str.replace('_', ' ').str.title(),
                          grouped['quantity'], label=bar_labels, color=colors)
            ax.set_ylabel(units_label, fontsize=14)
            ax.set_title(main_title, fontsize=16, fontweight="bold",
                         y=0.96, color="#2C3E50",
                         bbox=dict(boxstyle="round,pad=0.6", facecolor='#ECF0F1', alpha=0.8))
            ax.set_facecolor("#DBDBDB")
            ax.grid(axis='y', linestyle='--', alpha=0.7)
            ax.legend(title=legend_title, loc="center left",
                      bbox_to_anchor=(1, 0, 0.5, 1), fontsize=11)
            ax.set_ylim(0, max_quantity)

            # Anotar las barras usando el objeto bars
            for bar in bars:
                height = bar.get_height()
                ax.annotate(str(int(height)),
                            (bar.get_x() + bar.get_width() / 2., height),
                            ha='center', va='bottom', fontsize=12, color="black")

            # Ajustar layout específico para el gráfico de barras
            plt.subplots_adjust(top=0.88, bottom=0.12, left=0.10, right=0.75)
            plt.tight_layout(pad=3.0, rect=(0, 0, 1, 0.88))
        else:
            # Para gráfico de pie, usar layout estándar
            plt.tight_layout(pad=3.0)

        # Guardar gráfico si se creó una figura
        if fig:
            graphic_buffer = BytesIO()
            fig.savefig(graphic_buffer, format='png',
                        dpi=300, bbox_inches='tight')
            graphic_buffer.seek(0)
            graphic_content = graphic_buffer.getvalue()

            graphic_data_for_front = base64.b64encode(
                graphic_content).decode('utf-8')
            plt.close(fig)

    # Manejo de respuestas según parámetros
    if excel and graphic and graphic_content:
        # Excel con gráfico en segunda hoja
        excel_buffer = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws:
            ws.title = 'Movimientos Entrada vs Salida'

            column_mapping = {
                'id': 'ID', 'product__product_code': 'Código Producto',
                'product__name': 'Nombre Producto', 'product__description': 'Descripción',
                'product__brand': 'Marca', 'product__model': 'Modelo',
                'from_location__id': 'ID Ubicación Origen', 'from_location__name': 'Ubicación Origen',
                'to_location__id': 'ID Ubicación Destino', 'to_location__name': 'Ubicación Destino',
                'quantity': 'Cantidad', 'reason': 'Motivo', 'reference_type': 'Tipo Referencia',
                'reference_id': 'ID Referencia', 'occurred_at': 'Fecha de Ocurrencia'
            }
            ws = format_excel_worksheet(
                ws, df_movements, column_mapping, header_color="E74C3C")

            # Agregar segunda hoja con gráfico proporcional
            ws_graphic = wb.create_sheet(title="Gráfico")
            img_buffer = BytesIO(graphic_content)
            img = Image(img_buffer)
            # Tamaño proporcional para Excel
            img.width = 480
            img.height = 320
            ws_graphic.add_image(img, 'A1')

            wb.save(excel_buffer)
            excel_buffer.seek(0)

            if download_graphic:
                # ZIP con Excel y PNG
                zip_buffer = BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w') as zipf:
                    zipf.writestr("movements.xlsx", excel_buffer.getvalue())
                    zipf.writestr("graphic.png", graphic_content)
                zip_buffer.seek(0)
                resp = HttpResponse(zip_buffer.getvalue(),
                                    content_type='application/zip')
                resp['Content-Disposition'] = 'attachment; filename="movements_and_graphic.zip"'
                return resp
            else:
                # Solo Excel con gráfico
                resp = HttpResponse(excel_buffer.getvalue(
                ), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                resp['Content-Disposition'] = 'attachment; filename="movements.xlsx"'
                return resp

    elif excel and not graphic:
        # Solo Excel sin gráfico
        excel_buffer = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws:
            ws.title = 'Movimientos Entrada vs Salida'

            column_mapping = {
                'id': 'ID', 'product__product_code': 'Código Producto',
                'product__name': 'Nombre Producto', 'product__description': 'Descripción',
                'product__brand': 'Marca', 'product__model': 'Modelo',
                'from_location__id': 'ID Ubicación Origen', 'from_location__name': 'Ubicación Origen',
                'to_location__id': 'ID Ubicación Destino', 'to_location__name': 'Ubicación Destino',
                'quantity': 'Cantidad', 'reason': 'Motivo', 'reference_type': 'Tipo Referencia',
                'reference_id': 'ID Referencia', 'occurred_at': 'Fecha de Ocurrencia'
            }
            ws = format_excel_worksheet(
                ws, df_movements, column_mapping, header_color="E74C3C")

            wb.save(excel_buffer)
            excel_buffer.seek(0)
            resp = HttpResponse(excel_buffer.getvalue(
            ), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = 'attachment; filename="movements.xlsx"'
            return resp

    elif graphic and download_graphic and graphic_content:
        # Solo PNG para descarga (solo si graphic=True)
        resp = HttpResponse(graphic_content, content_type='image/png')
        resp['Content-Disposition'] = 'attachment; filename="input_vs_output_graphic.png"'
        return resp

    else:
        # JSON con gráfico base64 para front-end
        return JsonResponse({
            "success": True,
            "message": "Data retrieved successfully",
            "data": df_movements.to_dict(orient="records"),
            "graphic_base64": graphic_data_for_front
        })


# Resumen de Ventas
def sales_summary(
    from_date: date,
    to_date: date,
    month_compare: int = 2,
    language_graphic: str = 'es',
):
    """
    Genera un resumen de ventas en un rango de fechas.

    Calcula ingresos totales, unidades vendidas, productos más vendidos y
    genera dos gráficos (ingresos por mes y comparación con periodo previo).

    Args:
        from_date (date): Fecha inicial (inclusive).
        to_date (date, optional): Fecha final (inclusive). Por defecto hoy.
        month_compare (int, optional): Número de meses para comparar hacia atrás (usa 30 días por mes). Default 2.
        language_graphic (str, optional): 'es' o 'en' para títulos/etiquetas. Default 'es'.

    Returns:
        HttpResponse: Un archivo Excel (`resumen_ventas.xlsx`) que incluye los datos y gráficos.

    Raises:
        ValueError: Si las fechas son inválidas o `month_compare` no cumple restricciones.
    """
    try:

        if from_date < date.today()-timedelta(days=365*3):
            raise ValueError(
                "La fecha de inicio debe estar dentro de los últimos 3 años")
        if month_compare < 0:
            raise ValueError(
                "month_compare must be a positive integer or zero")
        if not isinstance(month_compare, int):
            raise ValueError("month_compare must be an integer")
        if month_compare >= 36:
            raise ValueError("month_compare must be less than 36")

        purchases_qs = Purchase.objects.filter(
            purchase_date__date__gte=from_date,
            purchase_date__date__lte=to_date,
        )

        if not purchases_qs.exists():
            # Obtener información adicional para ayudar al usuario
            total_purchases = Purchase.objects.count()

            if total_purchases == 0:
                # No hay compras en la base de datos
                return not_found_error_response(
                    message="No se encontraron compras en la base de datos. Por favor, cree algunas compras primero para generar el reporte de ventas.",
                    data={
                        "error_type": "no_data_in_database",
                        "total_purchases_in_db": 0,
                        "requested_range": {
                            "from_date": from_date.strftime("%Y-%m-%d"),
                            "to_date": to_date.strftime("%Y-%m-%d")
                        },
                        "suggestions": [
                            "Crear compras de prueba desde el panel de administración",
                            "Verificar que el sistema de compras esté funcionando correctamente",
                            "Contactar al administrador del sistema"
                        ]
                    }
                )
            else:
                # Hay compras en la DB pero no en el rango especificado
                # Obtener el rango de fechas de las compras existentes
                oldest_purchase = Purchase.objects.order_by(
                    'purchase_date').first()
                newest_purchase = Purchase.objects.order_by(
                    '-purchase_date').first()

                return not_found_error_response(
                    message=f"No se encontraron compras en el rango de fechas especificado ({from_date.strftime('%Y-%m-%d')} a {to_date.strftime('%Y-%m-%d')}). La base de datos contiene {total_purchases} compras en total.",
                    data={
                        "error_type": "no_data_in_range",
                        "total_purchases_in_db": total_purchases,
                        "requested_range": {
                            "from_date": from_date.strftime("%Y-%m-%d"),
                            "to_date": to_date.strftime("%Y-%m-%d")
                        },
                        "available_range": {
                            "oldest_purchase": oldest_purchase.purchase_date.date().strftime("%Y-%m-%d") if oldest_purchase else None,
                            "newest_purchase": newest_purchase.purchase_date.date().strftime("%Y-%m-%d") if newest_purchase else None
                        },
                        "suggestions": [
                            f"Ajustar el rango de fechas para incluir datos entre {oldest_purchase.purchase_date.date().strftime('%Y-%m-%d') if oldest_purchase else 'N/A'} y {newest_purchase.purchase_date.date().strftime('%Y-%m-%d') if newest_purchase else 'N/A'}",
                            "Crear compras adicionales en el rango de fechas deseado",
                            "Verificar que las fechas estén en formato correcto (YYYY-MM-DD)"
                        ]
                    }
                )

        df_purchases = pd.DataFrame.from_records(list(purchases_qs.values()))
        # Conservar la fecha original de purchase_date antes de convertir
        df_purchases['original_purchase_date'] = df_purchases['purchase_date']
        df_purchases['purchase_date'] = pd.to_datetime(
            df_purchases['purchase_date']).dt.date

        df_purchases['total_amount'] = _to_decimal_series(
            df_purchases.get('total_amount'))

        purchase_ids = purchases_qs.values_list('id', flat=True)
        details_qs = PurchaseDetail.objects.filter(
            purchase_id__in=purchase_ids
        ).values(
            "purchase_id",
            "product__name",
            "quantity",
            "unit_price_at_purchase",
            "subtotal",
            "product__brand",
            "product__model",
            "product__product_code",
        )
        df_details = pd.DataFrame.from_records(list(details_qs))

        # DataFrame con todos los detalles y datos de compra
        df_purchases_details = pd.merge(
            df_details, df_purchases, left_on='purchase_id', right_on='id', how='left')
        df_purchases_details.drop(columns=["id"], inplace=True)

        # Usar la fecha original de purchase_date para mostrar con hora, no la procesada
        df_purchases_details['purchase_date'] = df_purchases_details['original_purchase_date']

        total_revenue = df_purchases['total_amount'].sum()
        total_units_sold = df_details['quantity'].sum()

        product_sales = df_details.groupby('product__name').agg(
            total_quantity_sold=('quantity', 'sum'),
            total_revenue=('subtotal', 'sum')
        ).sort_values(by='total_quantity_sold', ascending=False)
        if 'total_revenue' in product_sales.columns and not product_sales['total_revenue'].empty:
            product_sales['total_revenue'] = _to_decimal_series(
                product_sales['total_revenue'])

        # Uso de 30 días para simplicidad
        compare_from_date = from_date - timedelta(days=month_compare * 30)
        last_period_purchases_qs = Purchase.objects.filter(
            purchase_date__date__gte=compare_from_date,
            purchase_date__date__lte=from_date
        )

        df_last_period_purchases = pd.DataFrame.from_records(
            list(last_period_purchases_qs.values()))
        if not df_last_period_purchases.empty:
            df_last_period_purchases['purchase_date'] = pd.to_datetime(
                df_last_period_purchases['purchase_date']).dt.date
            # Usar la misma función de conversión segura que el DataFrame principal
            df_last_period_purchases['total_amount'] = _to_decimal_series(
                df_last_period_purchases.get('total_amount'))

        # Ingresos mensuales del período actual
        df_purchases['year_month'] = pd.to_datetime(
            df_purchases['purchase_date']).dt.to_period('M')
        monthly_revenue_current = df_purchases.groupby('year_month')[
            'total_amount'].sum()

        # Ensure revenue series are numeric (floats) for plotting. They may be Decimal objects.
        def _series_to_float(series):
            if series is None:
                return series
            if getattr(series, 'empty', False):
                return series
            try:
                return series.apply(lambda v: float(v) if v is not None and not _pd.isna(v) else 0.0)
            except Exception:
                try:
                    return series.astype(float)
                except Exception:
                    return series

        monthly_revenue_current = _series_to_float(monthly_revenue_current)

        # Ingresos mensuales del período anterior
        if not df_last_period_purchases.empty:
            df_last_period_purchases['year_month'] = pd.to_datetime(
                df_last_period_purchases['purchase_date']).dt.to_period('M')
            monthly_revenue_last = df_last_period_purchases.groupby('year_month')[
                'total_amount'].sum()
            monthly_revenue_last = _series_to_float(monthly_revenue_last)
        else:
            monthly_revenue_last = pd.Series(dtype=float)

        main_title_entry = "Resumen de Ventas"
        y_label = "Ingresos ($)"
        x_label = "Mes"
        main_title_output_compared = 'Comparación de Ingresos Mensuales'
        keys_graphic_compared = ['Periodo Actual', 'Periodo Anterior']
        if language_graphic == 'en':
            main_title_entry = "Sales Summary"
            y_label = "Revenue ($)"
            x_label = "Month"
            main_title_output_compared = 'Monthly Revenue Comparison'
            keys_graphic_compared = ['Current Period', 'Previous Period']

        # Gráfico 1: Ingresos totales por mes
        fig_revenue, ax_revenue = plt.subplots(figsize=(10, 6))
        monthly_revenue_current.plot(
            kind='bar', ax=ax_revenue, color='skyblue', edgecolor='black')
        ax_revenue.set_title(main_title_entry, fontsize=12, fontweight="bold",
                             y=0.96, color="#2C3E50",
                             bbox=dict(boxstyle="round,pad=0.6", facecolor='#ECF0F1', alpha=0.8))
        ax_revenue.set_xlabel(x_label, fontsize=12)
        ax_revenue.set_ylabel(y_label, fontsize=12)
        ax_revenue.tick_params(axis='x', rotation=45)
        ax_revenue.set_facecolor("#DBDBDB")
        ax_revenue.grid(axis='y', linestyle='--', alpha=0.7)
        # Ajustar layout con espacio para el título
        plt.subplots_adjust(top=0.88, bottom=0.15, left=0.10, right=0.95)
        fig_revenue.tight_layout(pad=3.0, rect=(0, 0, 1, 0.88))
        img_revenue_buffer = io.BytesIO()
        fig_revenue.savefig(img_revenue_buffer, format='png')
        img_revenue_buffer.seek(0)
        plt.close(fig_revenue)

        # Gráfico 2: Comparación de ingresos
        fig_compare, ax_compare = plt.subplots(figsize=(10, 6))
        all_months = pd.concat([monthly_revenue_current, monthly_revenue_last],
                               axis=1, keys=keys_graphic_compared).fillna(0)
        all_months.plot(kind='bar', ax=ax_compare)
        ax_compare.set_title(main_title_output_compared, fontsize=12, fontweight="bold",
                             y=0.96, color="#2C3E50",
                             bbox=dict(boxstyle="round,pad=0.6", facecolor='#ECF0F1', alpha=0.8))
        ax_compare.set_xlabel(x_label, fontsize=12)
        ax_compare.set_ylabel(y_label, fontsize=12)
        ax_compare.tick_params(axis='x', rotation=45)
        ax_compare.set_facecolor("#DBDBDB")
        ax_compare.grid(axis='y', linestyle='--', alpha=0.7)
        ax_compare.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        # Ajustar layout con espacio para el título
        plt.subplots_adjust(top=0.88, bottom=0.15, left=0.10, right=0.85)
        fig_compare.tight_layout(pad=3.0, rect=(0, 0, 1, 0.88))
        img_compare_buffer = io.BytesIO()
        fig_compare.savefig(img_compare_buffer, format='png')
        img_compare_buffer.seek(0)
        plt.close(fig_compare)

        output = io.BytesIO()
        wb = openpyxl.Workbook()

        # Hoja 1: Datos de Compras
        ws_data = wb.create_sheet('Datos')
        if wb.worksheets and len(wb.worksheets) > 1:
            # Remover la hoja por defecto si creamos una nueva
            wb.remove(wb.worksheets[0])

        # Filtrar columnas de auditoría que no deben aparecer en reportes
        # Incluimos 'original_purchase_date' y 'created_at' en la exclusión, pero conservamos 'purchase_date'
        audit_columns = ['updated_at', 'created_by_id',
                         'updated_by_id', 'original_purchase_date', 'created_at']
        df_filtered = df_purchases_details.drop(
            columns=[col for col in audit_columns if col in df_purchases_details.columns], errors='ignore')

        # Mejorar nombres de cabeceras para hacerlas más legibles
        column_mapping = {
            'product__name': 'Producto',
            'product__brand': 'Marca',
            'product__model': 'Modelo',
            'product__product_code': 'Código',
            'quantity': 'Cantidad',
            'unit_price_at_purchase': 'Precio Unitario',
            'subtotal': 'Subtotal',
            'total_amount': 'Total',
            'status': 'Estado',
            'purchase_date': 'Fecha de Compra',
            'total_installments_count': 'Cuotas',
            'discount_applied': 'Descuento'
        }

        # Aplicar formato usando la función auxiliar
        ws_data = format_excel_worksheet(
            ws_data, df_filtered, column_mapping, header_color="366092")

        # Hoja 2: Resumen y Gráficos
        ws_summary = wb.create_sheet('Resumen y Graficos')

        # Escribir tabla de KPIs usando la función auxiliar
        summary_data = [
            ('Ingresos Totales', f"${total_revenue:,.2f}"),
            ('Volumen de Ventas', f"{total_units_sold:,}")
        ]
        ws_summary = format_summary_worksheet(
            ws_summary, summary_data, "Resumen de Ventas", "366092")

        # Insertar imágenes desde los búferes de memoria
        img_revenue = Image(img_revenue_buffer)
        # Tamaño proporcional para Excel
        img_revenue.width = 480
        img_revenue.height = 320
        ws_summary.add_image(img_revenue, 'A8')

        img_compare = Image(img_compare_buffer)
        # Tamaño proporcional para Excel
        img_compare.width = 480
        img_compare.height = 320
        ws_summary.add_image(img_compare, 'E8')

        # Guardar el libro de trabajo en el búfer de memoria
        wb.save(output)

        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=resumen_ventas.xlsx'

        return response
    except ValueError as ve:
        # Manejar errores de validación específicamente
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Error de validación en sales_summary: {str(ve)}")

    # Retornar error de validación con código 400
        return {
            'error': str(ve),
            'error_type': 'validation_error',
            'suggestions': [
                "Verifique que las fechas estén en formato YYYY-MM-DD",
                "Asegúrese de que from_date sea anterior a to_date",
                "Compruebe que las fechas estén dentro del rango permitido (últimos 5 años)",
                "Verifique que month_compare sea un número entero entre 0 y 35"
            ]
        }
    except Exception as e:
        # Registrar error completo en logs del servidor (no exponer al cliente)
        import traceback
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error en sales_summary: {str(e)}", exc_info=True)
        logger.error(f"Traceback completo: {traceback.format_exc()}")

    # Retornar error seguro sin exponer información sensible
    if "Row numbers must be between" in str(e):
        error_msg = "El dataset es demasiado grande para generar en Excel. Intente con un rango de fechas más pequeño."
    elif "memory" in str(e).lower():
        error_msg = "Insuficiente memoria para procesar el reporte. Reduzca el rango de fechas."
    else:
        error_msg = "Error interno al procesar el reporte. Contacte al administrador."

        return {'error': error_msg}


# productos mas vendidos
def most_sold_products(
    from_date: date,
    to_date: date = date.today(),
    limit: int = 10,
    excel: bool = False,
    graphic: bool = True,
    download_graphic: bool = False,
    language_graphic: str = 'es',
):
    """
    Devuelve los productos más y menos vendidos en un rango de fechas.

    Args:
        from_date (date): Fecha inicial (inclusive).
        to_date (date, optional): Fecha final (inclusive). Default hoy.
        limit (int, optional): Número máximo de productos a incluir en el ranking. Default 10.
        excel (bool, optional): Si generar archivo Excel. Default False.
        graphic (bool, optional): Si generar gráfico. Default True.
        download_graphic (bool, optional): Si descargar gráfico como PNG en ZIP. Default False.
        language_graphic (str, optional): Idioma del gráfico ('es' o 'en'). Default 'es'.

    Returns:
        dict | JsonResponse | HttpResponse: Si hay datos, devuelve un diccionario con claves:
            - most_sold_product
            - least_sold_product
            - percentage_most_sold
            - total_products_sold
            - total_revenue
            - detailed_data (lista)

        Si excel=True devuelve HttpResponse con archivo Excel.
        Si excel=True y download_graphic=True devuelve ZIP con Excel + PNG.
        Si no hay datos devuelve un `JsonResponse` con `success: False` y status 404.

    Raises:
        ValueError: Si las fechas son inconsistentes.
    """
    if from_date >= to_date:
        raise ValueError("from_date debe ser anterior a to_date")
    if to_date > date.today():
        raise ValueError("to_date no puede ser mayor que la fecha actual")
    if from_date > date.today():
        raise ValueError("from_date no puede ser mayor que la fecha actual")
    if from_date < date.today()-timedelta(days=365*3):
        raise ValueError("from_date must be within the last 3 years")

    details_qs = PurchaseDetail.objects.filter(
        purchase__created_at__date__gte=from_date,
        purchase__created_at__date__lte=to_date,
    ).values(
        "purchase_id",
        "product__name",
        "quantity",
        "subtotal",
        "product__brand",
        "product__model",
        "product__product_code",
    )

    df_details = pd.DataFrame.from_records(list(details_qs))

    if df_details.empty:
        # Verificar si hay compras pero sin detalles
        purchases_in_range = Purchase.objects.filter(
            created_at__date__gte=from_date,
            created_at__date__lte=to_date,
        ).count()

        total_purchases = Purchase.objects.count()
        total_details = PurchaseDetail.objects.count()

        if purchases_in_range > 0:
            # Hay compras pero sin detalles
            return JsonResponse({
                "success": False,
                "message": f"Se encontraron {purchases_in_range} compras en el rango especificado, pero no tienen detalles de productos asociados.",
                "data": {
                    "error_type": "purchases_without_details",
                    "purchases_in_range": purchases_in_range,
                    "total_purchases_in_db": total_purchases,
                    "total_details_in_db": total_details,
                    "requested_range": {
                        "from_date": from_date.strftime("%Y-%m-%d"),
                        "to_date": to_date.strftime("%Y-%m-%d")
                    },
                    "suggestions": [
                        "Verificar que las compras tengan productos asociados",
                        "Revisar la integridad de los datos de PurchaseDetail",
                        "Contactar al administrador del sistema"
                    ]
                }
            }, status=404)
        elif total_purchases == 0:
            # No hay compras en absoluto
            return JsonResponse({
                "success": False,
                "message": "No se encontraron compras en la base de datos. Cree algunas compras primero.",
                "data": {
                    "error_type": "no_purchases_in_database",
                    "total_purchases_in_db": 0,
                    "requested_range": {
                        "from_date": from_date.strftime("%Y-%m-%d"),
                        "to_date": to_date.strftime("%Y-%m-%d")
                    },
                    "suggestions": [
                        "Crear compras desde el panel de administración",
                        "Verificar la configuración del sistema de compras"
                    ]
                }
            }, status=404)
        else:
            # Hay compras en la DB pero no en el rango especificado
            oldest_purchase = Purchase.objects.order_by('created_at').first()
            newest_purchase = Purchase.objects.order_by('-created_at').first()

            return JsonResponse({
                "success": False,
                "message": f"No se encontraron detalles de compras en el rango especificado ({from_date.strftime('%Y-%m-%d')} a {to_date.strftime('%Y-%m-%d')}). La base de datos contiene {total_purchases} compras.",
                "data": {
                    "error_type": "no_purchase_details_in_range",
                    "total_purchases_in_db": total_purchases,
                    "total_details_in_db": total_details,
                    "requested_range": {
                        "from_date": from_date.strftime("%Y-%m-%d"),
                        "to_date": to_date.strftime("%Y-%m-%d")
                    },
                    "available_range": {
                        "oldest_purchase": oldest_purchase.created_at.date().strftime("%Y-%m-%d") if oldest_purchase else None,
                        "newest_purchase": newest_purchase.created_at.date().strftime("%Y-%m-%d") if newest_purchase else None
                    },
                    "suggestions": [
                        f"Ajustar el rango de fechas entre {oldest_purchase.created_at.date().strftime('%Y-%m-%d') if oldest_purchase else 'N/A'} y {newest_purchase.created_at.date().strftime('%Y-%m-%d') if newest_purchase else 'N/A'}",
                        "Crear más compras en el rango deseado"
                    ]
                }
            }, status=404)

    if 'subtotal' in df_details.columns and not df_details['subtotal'].empty:
        df_details['subtotal'] = _to_decimal_series(df_details['subtotal'])

    # Agrupar y ordenar productos por cantidad vendida
    grouped = df_details.groupby('product__name').agg(
        total_quantity_sold=('quantity', 'sum'),
        total_revenue=('subtotal', 'sum')
    ).sort_values(by='total_quantity_sold', ascending=False).reset_index()

    # Aseguramos que la columna de ingresos sea Decimal
    if 'total_revenue' in grouped.columns and not grouped['total_revenue'].empty:
        grouped['total_revenue'] = _to_decimal_series(grouped['total_revenue'])

    # Aplicar límite de productos
    grouped_limited = grouped.head(limit)

    # Generar gráfico si se solicita
    graphic_file = None
    graphic_data_for_front = None
    graphic_content = None
    if graphic:
        title = get_translation('most_sold_products_title', language_graphic)
        xlabel = get_translation('quantity_label', language_graphic) + \
            (" Vendida" if language_graphic == 'es' else " Sold")
        ylabel = get_translation('product_name', language_graphic)

        fig, ax = plt.subplots(figsize=(10, 5))

        # Crear gráfico de barras horizontal
        bars = ax.barh(range(len(grouped_limited)), grouped_limited['total_quantity_sold'],
                       color=['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                              '#8c564b', '#e377c2', '#7f7f7f', "#ebd495", '#17becf'][:len(grouped_limited)])

        # Configurar etiquetas
        ax.set_yticks(range(len(grouped_limited)))
        ax.set_yticklabels([name[:25] + '...' if len(name) > 25 else name
                           for name in grouped_limited['product__name']], fontsize=10)
        ax.set_xlabel(xlabel, fontsize=12)
        ax.set_ylabel(ylabel, fontsize=12)
        ax.set_title(title, fontsize=12, fontweight="bold",
                     y=0.96, color="#2C3E50",
                     bbox=dict(boxstyle="round,pad=0.6", facecolor='#ECF0F1', alpha=0.8))

        # Agregar valores en las barras
        for i, bar in enumerate(bars):
            width = bar.get_width()
            ax.text(width + max(grouped_limited['total_quantity_sold']) * 0.01,
                    bar.get_y() + bar.get_height()/2,
                    f'{int(width)}', ha='left', va='center', fontsize=9)

        # Ajustar diseño con espacio suficiente para el título
        plt.subplots_adjust(top=0.90, bottom=0.10, left=0.20, right=0.95)
        plt.tight_layout(pad=5.0, rect=(0, 0, 1, 0.90))

        # Guardar en buffer
        img_buffer = io.BytesIO()
        fig.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close(fig)

        # Guardar el contenido para reutilización
        graphic_content = img_buffer.getvalue()
        graphic_file = io.BytesIO(graphic_content)
        graphic_data_for_front = base64.b64encode(
            graphic_content).decode('utf-8')

    # Generar archivo Excel si se solicita
    excel_file = None
    if excel:
        output = io.BytesIO()
        wb = openpyxl.Workbook()

        # Crear hoja de datos
        ws_data = wb.create_sheet('Top Productos')
        if wb.worksheets and len(wb.worksheets) > 1:
            wb.remove(wb.worksheets[0])

        # Mejorar nombres de cabeceras
        column_mapping = {
            'product__name': 'Nombre Producto',
            'total_quantity_sold': 'Total Vendido',
            'total_revenue': 'Ingresos Totales'
        }

        # Aplicar formato
        ws_data = format_excel_worksheet(
            ws_data, grouped_limited, column_mapping, header_color="2E86AB")

        # Crear hoja de resumen
        ws_summary = wb.create_sheet('Resumen')
        summary_data = [
            ('Producto Más Vendido', f"{grouped.iloc[0]['product__name']}"),
            ('Cantidad Más Vendida',
             f"{grouped.iloc[0]['total_quantity_sold']:,}"),
            ('Total Productos Analizados', f"{len(grouped):,}"),
            ('Total Ingresos', f"${grouped['total_revenue'].sum():,.2f}")
        ]
        ws_summary = format_summary_worksheet(
            ws_summary, summary_data, f"Top {limit} Productos", "2E86AB")

        if graphic or download_graphic:
            img_chart = Image(graphic_file)
            # Tamaño proporcional para Excel
            img_chart.width = 480
            img_chart.height = 320
            ws_summary.add_image(img_chart, 'E5')

        wb.save(output)
        output.seek(0)
        excel_file = output

    # Convertir los DataFrames a diccionarios y luego aplicar conversión de tipos numpy
    most_sold_product_data = convert_numpy_types(
        grouped.head(1).to_dict(orient='records')[0])
    least_sold_product_data = convert_numpy_types(
        grouped.tail(1).to_dict(orient='records')[0]
    ) if len(grouped) > 1 else 'N/A'
    detailed_data = convert_numpy_types(
        grouped_limited.to_dict(orient='records'))

    # Convertir valores individuales
    total_quantity_sold = convert_numpy_types(
        grouped['total_quantity_sold'].sum())
    percentage_most_sold = convert_numpy_types(
        (grouped.iloc[0]['total_quantity_sold'] / total_quantity_sold * 100).round(2))
    total_revenue = convert_numpy_types(grouped['total_revenue'].sum())

    response_data = {
        'success': True,
        'message': 'Data retrieved successfully',
        'data': {
            'most_sold_product': most_sold_product_data,
            'least_sold_product': least_sold_product_data,
            'percentage_most_sold': f"{percentage_most_sold}%",
            'total_products_sold': total_quantity_sold,
            'total_revenue': f"${total_revenue:,.2f}",
            'detailed_data': detailed_data
        },
        'graphic_base64': graphic_data_for_front
    }

    # Manejo de respuestas según parámetros
    if not excel and graphic_content and download_graphic:
        # Solo PNG descargable
        response = HttpResponse(graphic_content, content_type='image/png')
        response['Content-Disposition'] = 'attachment; filename=grafico_top_productos.png'
        return response

    if excel_file and not graphic:
        # Solo Excel
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=top_productos_vendidos.xlsx'
        return response

    if excel_file and graphic_content and download_graphic:
        # ZIP con Excel + gráfico PNG
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            zip_file.writestr('top_productos_vendidos.xlsx',
                              excel_file.getvalue())
            zip_file.writestr('grafico_top_productos.png', graphic_content)

        zip_buffer.seek(0)
        response = HttpResponse(
            zip_buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=top_productos_completo.zip'
        return response

    if excel_file:
        # Solo Excel
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=top_productos_vendidos.xlsx'
        return response

    # Respuesta JSON (con o sin gráfico base64)
    return response_data


# Análisis de cuotas vencidas
def overdue_installments(
    from_date: date,
    to_date: date = date.today(),
    language_graphic: str = 'es',
    download_graphic: bool = False,
    output_format: str = 'json',
    excel: bool = False,
    graphic: bool = True,
):
    """
    Analiza cuotas vencidas con métricas avanzadas de aging y tendencias temporales.

    Calcula métricas tradicionales (mora, descuento, recargo) PLUS análisis avanzado:
    - Distribución por días de mora (1-30, 31-60, 61-90, 91-180, +180 días)
    - Impacto monetario por rango de aging 
    - Evolución temporal de vencimientos
    - Casos críticos (+90 días de mora)
    - Dashboard visual con 4 gráficos especializados

    Args:
        from_date (date): Fecha inicial (inclusive).
        to_date (date, optional): Fecha final (inclusive). Default hoy.
        language_graphic (str, optional): 'es' o 'en' para etiquetas. Default 'es'.
        download_graphic (bool, optional): Si descargar gráfico como PNG. Default False.
        output_format (str, optional): Formato de salida ('json', 'excel'). Default 'json'.
        excel (bool, optional): Si generar archivo Excel. Default False.
        graphic (bool, optional): Si generar dashboard de 4 gráficos analíticos. Default True.

    Returns:
        dict | JsonResponse | HttpResponse: JSON con métricas tradicionales + aging_analysis,
            HttpResponse con Excel, o ZIP según parámetros. El campo aging_analysis incluye:
            - average_days_overdue: Promedio de días de mora
            - maximum_days_overdue: Máximo días de mora en el dataset
            - critical_cases_90_plus: Casos con +90 días de mora
            - distribution_by_aging: Distribución por rangos de días
            - monetary_impact_by_aging: Impacto monetario por rango

    Raises:
        ValueError: Si las fechas son inválidas.
    """
    if from_date >= to_date:
        raise ValueError("from_date debe ser anterior a to_date")
    if to_date > date.today():
        raise ValueError("to_date no puede ser mayor que la fecha actual")
    if from_date > date.today():
        raise ValueError("from_date no puede ser mayor que la fecha actual")
    if from_date < date.today()-timedelta(days=365*3):
        raise ValueError("from_date must be within the last 3 years")

    # Installment.due_date is a DateField and the status field is named 'state'
    installment_qs = Installment.objects.filter(
        due_date__gte=from_date,
        due_date__lte=to_date,
        state=Installment.State.OVERDUE,
    )

    if not installment_qs.exists():
        return {
            "success": False,
            "message": "No overdue installments found for the specified date range.",
            "data": []
        }

    df_installment = pd.DataFrame.from_records(list(installment_qs.values()))
    if 'due_date' in df_installment.columns:
        df_installment['due_date'] = pd.to_datetime(
            df_installment['due_date']).dt.date

    # Prefer 'amount_due' column from Installment; fallback to 'amount' or zeros
    if 'amount_due' in df_installment.columns and not df_installment['amount_due'].empty:
        df_installment['amount'] = _to_decimal_series(
            df_installment['amount_due'])
    elif 'amount' in df_installment.columns and not df_installment['amount'].empty:
        df_installment['amount'] = _to_decimal_series(df_installment['amount'])
    else:
        df_installment['amount'] = _pd.Series(
            [_Decimal('0')] * len(df_installment))

    df_installment_ids = installment_qs.values_list('id', flat=True)

    audit_qs = InstallmentAuditLog.objects.filter(
        installment_id__in=df_installment_ids
    ).values(
        "installment_id",
        "updated_at",
        "reason",
        "delta_json"
    )
    df_audit = pd.DataFrame.from_records(list(audit_qs))
    if not df_audit.empty and 'updated_at' in df_audit.columns:
        df_audit['updated_at'] = pd.to_datetime(df_audit['updated_at'])

    # delta_json may already be a dict (JSONField) or a JSON string; parse defensively
    def _safe_parse_delta(x):
        if x is None or _pd.isna(x):
            return {}
        if isinstance(x, dict):
            return x
        if isinstance(x, str):
            try:
                return json.loads(x)
            except Exception:
                return {}
        # fallback
        return {}

    if not df_audit.empty and 'delta_json' in df_audit.columns:
        df_audit['delta_json'] = df_audit['delta_json'].apply(
            _safe_parse_delta)

    # Merge seguro de datos
    df_merged = pd.merge(
        df_installment, df_audit, left_on='id', right_on='installment_id', how='left')

    # Limpiar columnas duplicadas
    if 'installment_id' in df_merged.columns:
        df_merged.drop(columns=["installment_id"], inplace=True)

    # Asegurar que las columnas necesarias existen
    if 'updated_at' not in df_merged.columns:
        df_merged['updated_at'] = pd.NaT
    if 'delta_json' not in df_merged.columns:
        df_merged['delta_json'] = pd.Series(
            [{}] * len(df_merged), index=df_merged.index)

    # Aplicar parsing seguro a delta_json después del merge
    df_merged['delta_json'] = df_merged['delta_json'].apply(
        lambda x: x if isinstance(x, dict) else (
            {} if pd.isna(x) or x is None else {})
    )

    # Ordenar y obtener el registro más reciente por cuota
    df_merged.sort_values(by=['id', 'updated_at'], ascending=[
                          True, False], inplace=True)
    df_merged_unique = df_merged.drop_duplicates(
        subset=['id'], keep='first').copy()

    # Cálculos de contadores con manejo seguro
    def safe_get_nested(x, key, index=1, default_value=0):
        """Obtiene valores anidados de forma segura."""
        if not isinstance(x, dict):
            return default_value
        value = x.get(key)
        if value is None:
            return default_value
        if isinstance(value, (list, tuple)) and len(value) > index:
            return value[index]
        return default_value

    # Contadores mejorados con manejo de errores
    count_mora = df_merged_unique[
        df_merged_unique['delta_json'].apply(
            lambda x: isinstance(x, dict) and x.get(
                'mora') is not None and bool(x.get('mora'))
        )
    ].shape[0]

    count_discount = df_merged_unique[
        df_merged_unique['delta_json'].apply(
            lambda x: safe_get_nested(x, 'discount_pct', 1, 0) > 0
        )
    ].shape[0]

    count_surcharge = df_merged_unique[
        df_merged_unique['delta_json'].apply(
            lambda x: safe_get_nested(x, 'surcharge_pct', 1, 0) > 0
        )
    ].shape[0]

    count_both = df_merged_unique[
        (df_merged_unique['delta_json'].apply(lambda x: safe_get_nested(x, 'discount_pct', 1, 0) > 0)) &
        (df_merged_unique['delta_json'].apply(
            lambda x: safe_get_nested(x, 'surcharge_pct', 1, 0) > 0))
    ].shape[0]

    # totals used to compute 'no change' and for the summary table
    total_overdue = df_installment['amount'].sum(
    ) if 'amount' in df_installment.columns else 0
    count_overdue = int(df_installment.shape[0])

    count_no_change = count_overdue - \
        (count_mora + count_discount + count_surcharge - count_both)

    if language_graphic == 'es':
        main_title = "Distribución de Cuotas Vencidas por Estado"
        labels = ['Con Mora', 'Con Descuento', 'Con Recargo', 'Sin Cambios']
    else:  # English
        main_title = "Distribution of Overdue Installments by Status"
        labels = ['With Late Fee', 'With Discount',
                  'With Surcharge', 'No Changes']

    sizes = [count_mora, count_discount, count_surcharge, count_no_change]

    if sum(sizes) == 0:
        sizes = [1]
        labels = ['No Data']

    # Continuar con el resto del procesamiento...

    #  CALCULAR DÍAS DE MORA ANTES DE USARLOS EN MÉTRICAS
    # Calcular días de mora para cada cuota vencida
    today = date.today()
    df_installment['days_overdue'] = df_installment['due_date'].apply(
        lambda x: (today - x).days if pd.notna(x) else 0
    )

    # Definir rangos de mora
    bins = [0, 30, 60, 90, 180, float('inf')]
    labels_mora = [
        get_translation('aging_1_30', language_graphic),
        get_translation('aging_31_60', language_graphic),
        get_translation('aging_61_90', language_graphic),
        get_translation('aging_91_180', language_graphic),
        get_translation('aging_180_plus', language_graphic)
    ]

    df_installment['mora_range'] = pd.cut(
        df_installment['days_overdue'],
        bins=bins,
        labels=labels_mora,
        right=False
    )

    # Calcular métricas avanzadas de días de mora
    if not df_installment.empty:
        # Análisis por rangos de días de mora (conversión segura a float)
        avg_days_overdue = float(df_installment['days_overdue'].mean())
        max_days_overdue = int(df_installment['days_overdue'].max())
        critical_cases = len(
            df_installment[df_installment['days_overdue'] > 90])  # Más de 90 días

        # Distribución por rangos de mora
        mora_distribution = df_installment['mora_range'].value_counts(
        ).to_dict()

        # Impacto monetario por rango (conversión segura a float)
        monetary_impact = {}
        for k, v in df_installment.groupby('mora_range')['amount'].sum().items():
            # Convertir Decimal a float para evitar errores
            monetary_impact[str(k)] = float(v)
    else:
        avg_days_overdue = 0.0
        max_days_overdue = 0
        critical_cases = 0
        mora_distribution = {}
        monetary_impact = {}

    # Convertir contadores a tipos seguros para cálculos de porcentaje
    count_mora = int(count_mora)
    count_discount = int(count_discount)
    count_surcharge = int(count_surcharge)
    count_both = int(count_both)
    count_no_change = int(count_no_change)
    count_overdue = int(count_overdue)

    data = {
        # Convertir Decimal a float
        'total_overdue_amount': f"${float(total_overdue):,.2f}",
        'total_overdue_count': count_overdue,
        'overdue_with_mora': {
            'count': count_mora,
            'percentage': f"{(count_mora / count_overdue * 100):.2f}%" if count_overdue > 0 else "0%"
        },
        'overdue_with_discount': {
            'count': count_discount,
            'percentage': f"{(count_discount / count_overdue * 100):.2f}%" if count_overdue > 0 else "0%"
        },
        'overdue_with_surcharge': {
            'count': count_surcharge,
            'percentage': f"{(count_surcharge / count_overdue * 100):.2f}%" if count_overdue > 0 else "0%"
        },
        'overdue_no_changes': {
            'count': count_no_change,
            'percentage': f"{(count_no_change / count_overdue * 100):.2f}%" if count_overdue > 0 else "0%"
        },
        'overdue_with_both': {
            'count': count_both,
            'percentage': f"{(count_both / count_overdue * 100):.2f}%" if count_overdue > 0 else "0%"
        },
        # 🚨 NUEVAS MÉTRICAS MÁS ÚTILES
        'aging_analysis': {
            'average_days_overdue': round(avg_days_overdue, 1),
            'maximum_days_overdue': max_days_overdue,
            'critical_cases_90_plus': critical_cases,
            'distribution_by_aging': mora_distribution,
            'monetary_impact_by_aging': {k: f"${v:,.2f}" for k, v in monetary_impact.items()}
        },
        'detailed_data': convert_numpy_types(df_installment.to_dict(orient='records'))
    }

    # Configurar matplotlib
    _configure_professional_matplotlib()

    # Generar análisis temporal avanzado si se solicita
    graphic_content = None
    graphic_data_for_front = None
    if graphic and not df_installment.empty:
        # Crear figura con subplots para análisis más rico
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(15, 12))

        # Configurar paleta profesional
        professional_colors = _get_professional_color_palette(6)

    # Gráfico 1: Distribución por días de mora (ya calculado anteriormente)
        mora_counts = df_installment['mora_range'].value_counts().sort_index()
        mora_amounts = df_installment.groupby(
            'mora_range')['amount'].sum().sort_index()

        # Gráfico de barras para distribución por días de mora
        if not mora_counts.empty:
            bars1 = ax1.bar(range(len(mora_counts)), mora_counts.values,
                            color=professional_colors[:len(mora_counts)], alpha=0.8)
            ax1.set_xlabel(get_translation('aging_distribution', language_graphic).replace(
                'Distribución por ', ''), fontsize=10)
            ax1.set_ylabel(f"{get_translation('quantity_label', language_graphic)} de Cuotas" if language_graphic ==
                           'es' else f"{get_translation('quantity_label', language_graphic)} of Installments", fontsize=10)
            ax1.set_title(f"{get_translation('aging_distribution', language_graphic)}\n(Análisis de Criticidad)" if language_graphic == 'es' else f"{get_translation('aging_distribution', language_graphic)}\n(Criticality Analysis)",
                          fontsize=14, fontweight="bold", pad=20, color="#2C3E50",
                          bbox=dict(boxstyle="round,pad=0.5", facecolor='#F8F9FA', alpha=0.8))
            ax1.set_xticks(range(len(mora_counts)))
            ax1.set_xticklabels(mora_counts.index, rotation=45, ha='right')

            # Agregar valores en las barras
            for bar in bars1:
                height = bar.get_height()
                ax1.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                         f'{int(height)}', ha='center', va='bottom', fontsize=9)
        else:
            no_data_text = get_translation('no_data_single', language_graphic) + (
                " de mora" if language_graphic == 'es' else " for overdue")
            ax1.text(0.5, 0.5, no_data_text, ha='center',
                     va='center', transform=ax1.transAxes)
            title_text = f"{get_translation('aging_distribution', language_graphic)}\n(Sin datos)" if language_graphic == 'es' else f"{get_translation('aging_distribution', language_graphic)}\n(No data)"
            ax1.set_title(title_text,
                          fontsize=14, fontweight='bold', pad=20, color="#2C3E50")

    # Gráfico 2: Impacto monetario por rango de mora
        if not mora_amounts.empty:
            # Convertir Decimal a float para evitar errores en matplotlib
            mora_amounts_float = mora_amounts.apply(float)

            bars2 = ax2.bar(range(len(mora_amounts_float)), mora_amounts_float.values,
                            color=professional_colors[:len(mora_amounts_float)], alpha=0.8)
            ax2.set_xlabel(get_translation('aging_distribution', language_graphic).replace(
                'Distribución por ', ''), fontsize=10)
            ax2.set_ylabel(f"Monto Total ($)" if language_graphic ==
                           'es' else f"Total Amount ($)", fontsize=10)
            ax2.set_title(f"{get_translation('monetary_impact', language_graphic)}\n(Priorización de Cobranza)" if language_graphic == 'es' else f"{get_translation('monetary_impact', language_graphic)}\n(Collection Prioritization)",
                          fontsize=14, fontweight="bold", pad=20, color="#2C3E50",
                          bbox=dict(boxstyle="round,pad=0.5", facecolor='#F8F9FA', alpha=0.8))
            ax2.set_xticks(range(len(mora_amounts_float)))
            ax2.set_xticklabels(mora_amounts_float.index,
                                rotation=45, ha='right')

            # Formatear valores monetarios en las barras
            max_amount = float(mora_amounts_float.max()) if len(
                mora_amounts_float) > 0 else 1
            for bar in bars2:
                height = bar.get_height()
                ax2.text(bar.get_x() + bar.get_width()/2., height + max_amount * 0.01,
                         f'${height:,.0f}', ha='center', va='bottom', fontsize=9)
        else:
            ax2.text(0.5, 0.5, 'Sin datos monetarios', ha='center',
                     va='center', transform=ax2.transAxes)
            ax2.set_title('Impacto Monetario por Mora\n(Sin datos)',
                          fontsize=14, fontweight='bold', pad=20, color="#2C3E50")

    # Gráfico 3: Evolución temporal (últimos 6 meses por semanas)
        try:
            # Agrupar por semana de vencimiento
            df_installment['due_week'] = pd.to_datetime(
                df_installment['due_date']).dt.to_period('W')
            weekly_evolution = df_installment.groupby('due_week').agg({
                'id': 'count',  # cantidad de cuotas
                # monto total convertido a float
                'amount': lambda x: x.apply(float).sum()
            }).tail(12)  # últimas 12 semanas

            if not weekly_evolution.empty:
                # Crear eje secundario para monto
                ax3_twin = ax3.twinx()

                # Línea para cantidad
                line1 = ax3.plot(range(len(weekly_evolution)), weekly_evolution['id'],
                                 color=professional_colors[0], marker='o', linewidth=2,
                                 label='Cantidad de Cuotas')
                ax3.set_xlabel('Semanas (Fecha de Vencimiento)', fontsize=10)
                ax3.set_ylabel('Cantidad de Cuotas', fontsize=10,
                               color=professional_colors[0])
                ax3.tick_params(axis='y', labelcolor=professional_colors[0])

                # Línea para monto en eje secundario
                line2 = ax3_twin.plot(range(len(weekly_evolution)), weekly_evolution['amount'],
                                      color=professional_colors[1], marker='s', linewidth=2,
                                      label='Monto Total ($)')
                ax3_twin.set_ylabel('Monto Total ($)',
                                    fontsize=10, color=professional_colors[1])
                ax3_twin.tick_params(
                    axis='y', labelcolor=professional_colors[1])

                ax3.set_title('Tendencia de Cuotas Vencidas\n(Últimas 12 Semanas)',
                              fontsize=14, fontweight="bold", pad=20, color="#2C3E50",
                              bbox=dict(boxstyle="round,pad=0.5", facecolor='#F8F9FA', alpha=0.8))
                # Mostrar cada 2 semanas
                ax3.set_xticks(range(0, len(weekly_evolution), 2))
                ax3.set_xticklabels(
                    [str(w) for w in weekly_evolution.index[::2]], rotation=45, ha='right')
            else:
                ax3.text(0.5, 0.5, 'Sin datos temporales', ha='center',
                         va='center', transform=ax3.transAxes)
                ax3.set_title(
                    'Tendencia de Cuotas Vencidas\n(Sin datos)', fontsize=14, fontweight='bold', pad=20, color="#2C3E50")
        except Exception as e:
            ax3.text(0.5, 0.5, f'Error en evolución temporal',
                     ha='center', va='center', transform=ax3.transAxes)
            ax3.set_title('Tendencia de Cuotas Vencidas\n(Error)',
                          fontsize=14, fontweight='bold', pad=20, color="#2C3E50")

        # GRÁFICO 4: Distribución actual por estado (versión compacta)
        pie_sizes = [count_mora, count_discount,
                     count_surcharge, count_no_change]
        pie_labels = labels[:4]

        if sum(pie_sizes) > 0:
            explode_values = [0.08, 0.04, 0.04, 0.02]
            wedges, texts = ax4.pie(pie_sizes, labels=None, autopct=None,
                                    colors=professional_colors[:4], startangle=90,
                                    explode=explode_values[:len(pie_sizes)],
                                    shadow=True,
                                    wedgeprops=dict(
                                        width=0.8, edgecolor='white', linewidth=1))

            # Crear leyenda con porcentajes
            total_counts = sum(pie_sizes)
            legend_labels = []
            for i, (label, count) in enumerate(zip(pie_labels, pie_sizes)):
                percentage = (count / total_counts) * \
                    100 if total_counts > 0 else 0
                legend_labels.append(f'{label}: {count} ({percentage:.1f}%)')

            legend_title = 'Distribución por Estado' if language_graphic == 'es' else 'Distribution by Status'
            ax4.legend(
                wedges,
                legend_labels,
                title=legend_title,
                loc="center left",
                bbox_to_anchor=(1, 0, 0.5, 1),
                fontsize=8,
                title_fontsize=9
            )

            ax4.set_title('Estados Actuales\n(Referencia)',
                          fontsize=14, fontweight="bold", pad=20, color="#2C3E50",
                          bbox=dict(boxstyle="round,pad=0.5", facecolor='#F8F9FA', alpha=0.8))
        else:
            ax4.text(0.5, 0.5, 'Sin Datos', ha='center',
                     va='center', transform=ax4.transAxes)
            ax4.set_title('Estados Actuales\n(Sin Datos)',
                          fontsize=14, fontweight='bold', pad=20, color="#2C3E50")

        # Ajustar layout con espacio suficiente para el título
        plt.subplots_adjust(top=0.92, bottom=0.10, left=0.08,
                            right=0.95, hspace=0.4, wspace=0.4)
        plt.tight_layout(pad=5.0, rect=(0, 0, 1, 0.92))

        # Agregar título general
        fig.suptitle(get_translation('overdue_analysis_title', language_graphic),
                     fontsize=16, fontweight="bold",
                     y=0.96, color="#2C3E50",
                     bbox=dict(boxstyle="round,pad=0.6", facecolor='#ECF0F1', alpha=0.8))

        # Guardar gráfico
        buffer = BytesIO()
        fig.savefig(buffer, format='png', dpi=300, bbox_inches='tight',
                    facecolor='white', edgecolor='none')
        buffer.seek(0)
        graphic_content = buffer.getvalue()
        plt.close(fig)

        # Para front-end
        import base64
        graphic_data_for_front = base64.b64encode(
            graphic_content).decode('utf-8')

    # Limpiar datos para Excel (eliminar objetos complejos)
    # Usar df_installment original para Excel
    df_for_excel = df_installment.copy()

    # Eliminar cualquier columna que pueda haber sido agregada durante el merge
    columns_to_remove = []
    for col in df_for_excel.columns:
        try:
            if not df_for_excel[col].empty:
                # Verificar múltiples muestras para ser más preciso
                sample_values = df_for_excel[col].dropna().head(5)
                for sample_value in sample_values:
                    # Verificar objetos complejos
                    if isinstance(sample_value, (dict, list)):
                        columns_to_remove.append(col)
                        break
                    # Verificar strings JSON problemáticos
                    elif isinstance(sample_value, str):
                        str_val = str(sample_value).strip()
                        if (str_val.startswith(('{', '[')) or
                            'new_value' in str_val or
                            'old_value' in str_val or
                            'timestamp' in str_val or
                            'field_changed' in str_val or
                                str_val.count("'") > 4):  # Indicador de dict convertido a string
                            columns_to_remove.append(col)
                            break
        except Exception:
            # Si hay cualquier problema al verificar, eliminar la columna
            columns_to_remove.append(col)

    # Eliminar columnas problemáticas únicas
    columns_to_remove = list(set(columns_to_remove))
    if columns_to_remove:
        df_for_excel = df_for_excel.drop(columns=columns_to_remove)
        # Para debugging
        print(f"DEBUG: Removed problematic columns: {columns_to_remove}")

    # Convertir fechas y decimales apropiadamente
    for col in df_for_excel.columns:
        try:
            if 'date' in col.lower() and not df_for_excel[col].empty:
                df_for_excel[col] = pd.to_datetime(
                    df_for_excel[col], errors='coerce').dt.strftime('%Y-%m-%d')
            elif df_for_excel[col].dtype == 'object':
                # Conversión segura a string
                df_for_excel[col] = df_for_excel[col].astype(str)
        except Exception:
            # Si hay error, convertir a string y limpiar
            try:
                df_for_excel[col] = df_for_excel[col].astype(str)
                # Limpiar cualquier residuo JSON
                df_for_excel[col] = df_for_excel[col].str.replace(
                    r'\{.*\}', '', regex=True)
                df_for_excel[col] = df_for_excel[col].str.replace(
                    r'\[.*\]', '', regex=True)
            except Exception:
                # Último recurso: llenar con valores vacíos
                df_for_excel[col] = ""
            df_for_excel[col] = df_for_excel[col].astype(str)

    # Generar Excel si se solicita
    excel_file = None
    if excel or output_format == 'excel':
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as OpenpyxlImage

        wb = Workbook()

    # Hoja de datos principales
        ws_data = wb.active
        if ws_data is not None:
            ws_data.title = 'Cuotas Vencidas'
        else:
            ws_data = wb.create_sheet('Cuotas Vencidas')

        # Mapeo profesional de columnas
        column_mapping = {
            'id': 'ID Cuota',
            'amount': 'Monto',
            'due_date': 'Fecha Vencimiento',
            'state': 'Estado',
            'payment_id': 'ID Pago',
            'created_at': 'Fecha Creación',
            'updated_at': 'Fecha Actualización'
        }

        ws_data = format_excel_worksheet(
            ws_data, df_for_excel, column_mapping,
            header_color="C0392B", title_color="FFFFFF"
        )

    # Hoja de resumen
        ws_summary = wb.create_sheet(title="Resumen Ejecutivo")
        summary_data = [
            ('Total Cuotas Vencidas', f'{count_overdue:,}'),
            # Convertir Decimal a float
            ('Monto Total Vencido', f'${float(total_overdue):,.2f}'),
            ('Cuotas con Mora',
             f'{count_mora:,} ({count_mora/count_overdue*100:.1f}%)' if count_overdue > 0 else '0'),
            ('Cuotas con Descuento',
             f'{count_discount:,} ({count_discount/count_overdue*100:.1f}%)' if count_overdue > 0 else '0'),
            ('Cuotas con Recargo',
             f'{count_surcharge:,} ({count_surcharge/count_overdue*100:.1f}%)' if count_overdue > 0 else '0'),
            ('Sin Cambios',
             f'{count_no_change:,} ({count_no_change/count_overdue*100:.1f}%)' if count_overdue > 0 else '0'),
        ]

        ws_summary = format_summary_worksheet(
            ws_summary, summary_data,
            title="Análisis de Cuotas Vencidas",
            header_color="C0392B"
        )

    # Agregar gráfico a Excel con tamaño proporcional
        if graphic_content:
            ws_chart = wb.create_sheet(title="Gráficos")
            img_buffer = BytesIO(graphic_content)
            img = OpenpyxlImage(img_buffer)
            # Tamaño más proporcional para Excel
            img.width = 480  # Reducido de 600 para mejor proporción
            img.height = 320  # Reducido de 400 para mejor proporción
            ws_chart.add_image(img, 'B2')

            # Título del gráfico
            ws_chart['B1'] = 'Distribución de Cuotas Vencidas'
            ws_chart['B1'].font = Font(bold=True, size=14)

        # Guardar Excel
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        excel_file = output

    # Manejo de respuestas según formato
    if excel_file and graphic_content and download_graphic:
        # ZIP con Excel + PNG
        import zipfile
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.writestr('cuotas_vencidas.xlsx', excel_file.getvalue())
            zipf.writestr('grafico_cuotas_vencidas.png', graphic_content)
        zip_buffer.seek(0)

        response = HttpResponse(zip_buffer.getvalue(),
                                content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=cuotas_vencidas_completo.zip'
        return response

    elif excel_file:
        # Solo Excel
        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=cuotas_vencidas.xlsx'
        return response

    elif graphic_content and download_graphic and not excel:
        # Solo PNG
        response = HttpResponse(graphic_content, content_type='image/png')
        response['Content-Disposition'] = 'attachment; filename=cuotas_vencidas.png'
        return response

    # Respuesta JSON con datos limpios
    data['detailed_data'] = convert_numpy_types(
        df_for_excel.to_dict(orient='records'))
    data['graphic_base64'] = graphic_data_for_front

    return {"success": True,
            "message": "Data retrieved successfully",
            "data": data}

# Análisis de métodos de pago


def payment_methods_analysis(
    from_date: date,
    to_date: date = date.today(),
    excel: bool = False,
    download_graphic: bool = False,
    language_graphic: str = 'es',
):
    """
    Analiza la distribución de métodos de pago en el periodo indicado.

    Agrupa pagos por `payment_method`, calcula totales y genera un gráfico tipo
    pastel que puede ser descargado como PNG o en Excel.

    Args:
        from_date (date): Fecha inicial (inclusive).
        to_date (date, optional): Fecha final (inclusive). Default hoy.
        excel (bool, optional): Si devuelve un Excel con datos y gráfico. Default False.
        download_graphic (bool, optional): Si descargar gráfico como PNG junto con Excel. Default False.
        language_graphic (str, optional): 'es' o 'en'. Default 'es'.

    Returns:
        dict | HttpResponse: Diccionario con `summary`, `total_amount` y `total_count`,
            `HttpResponse` con PNG si `download_graphic` True y no `excel`,
            `HttpResponse` con Excel si `excel` True,
            o `HttpResponse` con ZIP si `excel` True y `download_graphic` True.

    Raises:
        ValueError: Si las fechas son inválidas.
    """
    if from_date >= to_date:
        raise ValueError("from_date debe ser anterior a to_date")
    if to_date > date.today():
        raise ValueError("to_date no puede ser mayor que la fecha actual")
    if from_date > date.today():
        raise ValueError("from_date no puede ser mayor que la fecha actual")
    if from_date < date.today()-timedelta(days=365*3):
        raise ValueError("from_date debe ser dentro de los últimos 3 años")

    payment_qs = Payment.objects.filter(
        payment_date__date__gte=from_date,
        payment_date__date__lte=to_date)
    if not payment_qs.exists():
        return {
            "success": False,
            "message": "No se encontraron pagos para el rango de fechas especificado.",
            "data": []
        }
    df_payment = pd.DataFrame.from_records(list(payment_qs.values()))
    if 'payment_date' in df_payment.columns:
        df_payment['payment_date'] = pd.to_datetime(
            df_payment['payment_date']).dt.date
    if 'amount' in df_payment.columns and not df_payment['amount'].empty:
        df_payment['amount'] = _to_decimal_series(df_payment['amount'])
    if 'payment_method' in df_payment.columns:
        df_payment['payment_method'] = df_payment['payment_method'].astype(str)

    grouped_methods = df_payment.groupby(['payment_method'])
    grouped_payments_methods = grouped_methods.size().reset_index(
        name='counts').sort_values(by='counts', ascending=False)

    main_title = get_translation('payment_methods_title', language_graphic)
    legend_title = get_translation('payment_method', language_graphic)
    units_label = get_translation('total_units', language_graphic)

    fig, ax = plt.subplots(figsize=(10, 8), subplot_kw=dict(aspect="equal"))

    colors = ["#4E79A7", "#F28E2B", "#E15759"]

    explode = [0.1] + [0.05] * (len(grouped_payments_methods) - 1)

    pie_result = ax.pie(
        grouped_payments_methods['counts'],
        textprops=dict(color="black", fontsize=10, weight='bold'),
        colors=colors,
        startangle=90,
        explode=explode,
        shadow=True,
        wedgeprops=dict(width=0.9, edgecolor='white', linewidth=2),
        radius=1.1
    )

    # Handle pie result unpacking
    if len(pie_result) >= 1:
        wedges = pie_result[0]
    else:
        wedges = pie_result

    # Create legend with percentages
    total_counts = grouped_payments_methods['counts'].sum()
    legend_labels = []
    for i, (method, count) in enumerate(zip(grouped_payments_methods['payment_method'], grouped_payments_methods['counts'])):
        percentage = (count / total_counts) * 100
        clean_method = method.replace('_', ' ').title()
        legend_labels.append(
            f"{clean_method}: {percentage:.1f}% ({count:,} {units_label})")

    ax.legend(
        wedges,
        legend_labels,
        title=legend_title,
        loc="center left",
        bbox_to_anchor=(1, 0, 0.5, 1),
        fontsize=10
    )

    # Título más compacto y mejor posicionado
    compact_title = main_title.replace(
        ' de ', ' de\n') if language_graphic == 'es' else main_title.replace(' ', '\n', 1)

    ax.set_title(compact_title, fontsize=22, fontweight="bold",
                 y=0.96, color="#2C3E50",
                 bbox=dict(boxstyle="round,pad=0.6", facecolor='#ECF0F1', alpha=0.8))

    # Ajustar layout con espacio para el título
    plt.subplots_adjust(top=0.85, bottom=0.10, left=0.10, right=0.80)

    img_buffer = io.BytesIO()
    fig.savefig(img_buffer, format='png', bbox_inches='tight')
    img_buffer.seek(0)
    plt.close(fig)

    # Guardar el contenido del buffer para reutilización
    graphic_content = img_buffer.getvalue()
    graphic_file = io.BytesIO(graphic_content)

    # Preparar gráfico para front como base64
    import base64
    graphic_data_for_front = base64.b64encode(graphic_content).decode('utf-8')

    excel_file = None
    if excel:
        output = io.BytesIO()
        wb = openpyxl.Workbook()

        # Crear hoja de datos
        ws_data = wb.create_sheet('Datos')
        if wb.worksheets and len(wb.worksheets) > 1:
            # Remover la hoja por defecto si creamos una nueva
            wb.remove(wb.worksheets[0])

        # Filtrar columnas de auditoría
        audit_columns = ['updated_at', 'created_by_id',
                         'updated_by_id', 'created_at']
        df_filtered = df_payment.drop(
            columns=[col for col in audit_columns if col in df_payment.columns], errors='ignore')

        # Mejorar nombres de cabeceras
        column_mapping = {
            'payment_method': 'Método de Pago',
            'amount': 'Monto',
            'payment_date': 'Fecha de Pago',
            'external_ref': 'Referencia Externa',
            'installment_id': 'ID Cuota'
        }

        # Aplicar formato usando la función auxiliar
        ws_data = format_excel_worksheet(
            ws_data, df_filtered, column_mapping, header_color="27AE60")

        ws_summary = wb.create_sheet('Resumen y Graficos')

        # Usar la función auxiliar para el resumen
        summary_data = [
            ('Total Pagos', f"{grouped_payments_methods['counts'].sum():,}"),
            ('Total Monto Pagado', f"${df_payment['amount'].sum():,.2f}")
        ]
        ws_summary = format_summary_worksheet(
            ws_summary, summary_data, "Análisis de Métodos de Pago", "27AE60")

        img_pie = Image(graphic_file)
        # Tamaño proporcional para Excel
        img_pie.width = 480
        img_pie.height = 320
        ws_summary.add_image(img_pie, 'E5')

        wb.save(output)
        output.seek(0)
        excel_file = output

    # Preparar respuesta JSON con datos
    total_amount = convert_numpy_types(df_payment['amount'].sum())
    total_count = convert_numpy_types(grouped_payments_methods['counts'].sum())
    summary_data = convert_numpy_types(
        grouped_payments_methods.to_dict(orient='records'))

    response_data = {
        "success": True,
        "message": "Análisis de métodos de pago generado con éxito.",
        "data": {
            "summary": summary_data,
            "total_amount": f"${total_amount:,.2f}",
            "total_count": total_count
        },
        "graphic_base64": graphic_data_for_front
    }

    # Respuesta según formato solicitado
    # Solo Excel
    if excel_file and not download_graphic:
        resp = HttpResponse(excel_file.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="analisis_metodos_pago.xlsx"'
        return resp

    # Excel + gráfico PNG en ZIP
    if excel_file and download_graphic:
        from zipfile import ZipFile
        zip_buffer = BytesIO()
        with ZipFile(zip_buffer, 'w') as zipf:
            zipf.writestr("analisis_metodos_pago.xlsx", excel_file.getvalue())
            zipf.writestr("payment_methods_graphic.png", graphic_content)
        zip_buffer.seek(0)
        resp = HttpResponse(zip_buffer.getvalue(),
                            content_type='application/zip')
        resp['Content-Disposition'] = 'attachment; filename="payment_methods_analysis.zip"'
        return resp

    # Solo gráfico PNG para descarga
    if not excel_file and download_graphic:
        resp = HttpResponse(graphic_content, content_type='image/png')
        resp['Content-Disposition'] = 'attachment; filename="payment_methods_graphic.png"'
        return resp

    # Respuesta JSON normal
    from django.http import JsonResponse
    return JsonResponse(response_data)
